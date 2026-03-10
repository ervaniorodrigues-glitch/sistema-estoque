"""
Sistema de Controle de Estoque - Versão Web
Backend Flask com SQLite
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import unicodedata

# Carregar variáveis de ambiente do arquivo .env (se existir)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv não instalado, usar apenas variáveis de ambiente do sistema

# Função auxiliar para remover acentos
def remover_acentos(texto):
    """Remove acentos de uma string para facilitar buscas"""
    if not texto:
        return ''
    # Normaliza para NFD (separa caracteres base de acentos)
    nfkd = unicodedata.normalize('NFD', str(texto))
    # Remove os acentos (categoria Mn = Nonspacing Mark)
    return ''.join([c for c in nfkd if not unicodedata.combining(c)])

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'K8x#mP9$vL2@nQ5&wR7!jT4%yU6^bN3*cM1+dF0-eG8=')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Configuração do SQLite com otimizações
import os

# Configuração do banco de dados
# Se DATABASE_URL estiver definida (PostgreSQL no Render), usa ela
# Caso contrário, usa SQLite local
database_url = os.environ.get('DATABASE_URL')

if database_url:
    # PostgreSQL no Render ou outro servidor
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # SQLite local para desenvolvimento
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estoque.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}?check_same_thread=False'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'timeout': 15},
    'pool_pre_ping': True,
    'pool_recycle': 3600,
}
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static/uploads/produtos')

# Criar pasta se não existir
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.config['JSON_SORT_KEYS'] = False

# Configuração de sessão
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', 'False') == 'True'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400

db = SQLAlchemy(app)

# Middleware para desabilitar cache
@app.after_request
def disable_cache(response):
    """Desabilita cache para evitar lentidão e garantir que alterações sejam refletidas"""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Otimizações do SQLite para melhor performance (executar uma única vez)
def optimize_sqlite_once():
    """Ativa WAL mode e otimizações de performance no SQLite"""
    try:
        from sqlalchemy import text
        with app.app_context():
            db.session.execute(text('PRAGMA journal_mode=WAL'))
            db.session.execute(text('PRAGMA synchronous=NORMAL'))
            db.session.execute(text('PRAGMA cache_size=10000'))
            db.session.execute(text('PRAGMA temp_store=MEMORY'))
            db.session.commit()
    except Exception as e:
        print(f"Aviso: Não foi possível otimizar SQLite: {e}")

# Ativar otimizações imediatamente
optimize_sqlite_once()

# ==================== MODELS ====================

class Produto(db.Model):
    __tablename__ = 'produtos'
    
    codigo = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False, index=True)
    codigo_barras = db.Column(db.String(100), index=True)
    unidade = db.Column(db.String(10))
    marca = db.Column(db.String(100))
    categoria = db.Column(db.String(100), index=True)
    preco_compra = db.Column(db.Float, default=0)
    preco_venda = db.Column(db.Float, default=0)
    estoque_minimo = db.Column(db.Integer, default=0)
    estoque_maximo = db.Column(db.Integer, default=0)
    estoque_atual = db.Column(db.Integer, default=0)
    foto = db.Column(db.String(200))
    fornecedor = db.Column(db.String(200), index=True)
    observacao = db.Column(db.Text)
    ativo = db.Column(db.Boolean, default=True, index=True)
    data_cadastro = db.Column(db.DateTime, default=datetime.now)

class Funcionario(db.Model):
    __tablename__ = 'funcionarios'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False, index=True)
    cargo = db.Column(db.String(100))
    telefone = db.Column(db.String(20))
    celular = db.Column(db.String(20))
    email = db.Column(db.String(100), index=True)
    endereco = db.Column(db.String(200))
    cidade = db.Column(db.String(100))
    uf = db.Column(db.String(2))
    cep = db.Column(db.String(10))
    cpf = db.Column(db.String(14), index=True)
    data_admissao = db.Column(db.Date)
    salario = db.Column(db.Float)
    ativo = db.Column(db.Boolean, default=True, index=True)
    data_cadastro = db.Column(db.DateTime, default=datetime.now)

class Fornecedor(db.Model):
    __tablename__ = 'fornecedores'
    
    codigo = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False, index=True)
    telefone = db.Column(db.String(20))
    celular = db.Column(db.String(20))
    email = db.Column(db.String(100), index=True)
    categoria = db.Column(db.String(100), index=True)
    endereco = db.Column(db.String(200))
    cidade = db.Column(db.String(100))
    uf = db.Column(db.String(2))
    cep = db.Column(db.String(10))
    cnpj = db.Column(db.String(18), index=True)
    contato = db.Column(db.String(100))
    ativo = db.Column(db.Boolean, default=True, index=True)
    data_cadastro = db.Column(db.DateTime, default=datetime.now)

class Cliente(db.Model):
    __tablename__ = 'clientes'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False, index=True)
    email = db.Column(db.String(100), index=True)
    cpfcnpj = db.Column(db.String(18), unique=True, index=True)
    telefone = db.Column(db.String(20))
    celular = db.Column(db.String(20))
    endereco = db.Column(db.String(200))
    cidade = db.Column(db.String(100))
    uf = db.Column(db.String(2))
    cep = db.Column(db.String(10))
    anotacoes = db.Column(db.Text)
    ativo = db.Column(db.Boolean, default=True, index=True)
    data_cadastro = db.Column(db.DateTime, default=datetime.now)

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    usuario = db.Column(db.String(50), unique=True, nullable=False, index=True)
    senha = db.Column(db.String(200), nullable=False)
    admin = db.Column(db.Boolean, default=False)
    tipo = db.Column(db.String(20), default='comum')  # comum, intermediario, master
    ativo = db.Column(db.Boolean, default=True, index=True)
    data_cadastro = db.Column(db.DateTime, default=datetime.now)

class Unidade(db.Model):
    __tablename__ = 'unidades_div'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False, unique=True)

class Marca(db.Model):
    __tablename__ = 'marcas_div'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False, unique=True)

class Categoria(db.Model):
    __tablename__ = 'categorias_div'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False, unique=True)

class Operacao(db.Model):
    __tablename__ = 'operacoes_div'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False, unique=True)

class Ano(db.Model):
    __tablename__ = 'anos_div'
    
    id = db.Column(db.Integer, primary_key=True)
    ano = db.Column(db.Integer, nullable=False, unique=True)

class Emprestimo(db.Model):
    __tablename__ = 'emprestimos'
    
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    data = db.Column(db.String(20), nullable=False)
    funcionario = db.Column(db.String(200), nullable=False)
    cargo = db.Column(db.String(100))
    cod_produto = db.Column(db.Integer)
    descricao_item = db.Column(db.String(200), nullable=False)
    quantidade = db.Column(db.Integer, default=1)
    observacoes = db.Column(db.Text)
    status = db.Column(db.String(20), default='EMPRESTADO')
    data_devolucao_prevista = db.Column(db.String(20))
    data_devolucao = db.Column(db.String(20))
    data_cadastro = db.Column(db.DateTime, default=datetime.now)

# ==================== ROTAS ====================

@app.route('/init-database-render')
def init_database_render():
    """Rota especial para inicializar banco de dados no Render"""
    try:
        # Criar todas as tabelas
        db.create_all()
        
        # Verificar se usuário master já existe
        master = Usuario.query.filter_by(usuario='master').first()
        
        if not master:
            # Criar usuário master
            master = Usuario(
                nome='Administrador Master',
                usuario='master',
                senha=generate_password_hash('@Senha01'),
                admin=True,
                tipo='admin',
                ativo=True
            )
            db.session.add(master)
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Banco inicializado! Usuário master criado.',
                'usuario': 'master',
                'senha': '@Senha01'
            })
        else:
            return jsonify({
                'success': True,
                'message': 'Banco já está inicializado!',
                'usuario': 'master',
                'senha': '@Senha01'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao inicializar banco: {str(e)}'
        }), 500

@app.route('/')
def index():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            data = request.json
            if not data or 'usuario' not in data or 'senha' not in data:
                return jsonify({'success': False, 'message': 'Dados inválidos'})
            
            # Buscar usuário ignorando maiúsculas/minúsculas
            usuario = Usuario.query.filter(db.func.lower(Usuario.usuario) == db.func.lower(data['usuario'])).first()
            
            if not usuario:
                return jsonify({'success': False, 'message': 'Usuário ou senha inválidos'})
            
            senha_correta = check_password_hash(usuario.senha, data['senha'])
            
            if usuario and senha_correta:
                session.permanent = True
                session['usuario_id'] = usuario.id
                session['usuario_nome'] = usuario.nome
                session['usuario_admin'] = usuario.admin
                session['usuario_tipo'] = usuario.tipo if hasattr(usuario, 'tipo') else ('master' if usuario.admin else 'intermediario')
                return jsonify({'success': True})
            
            return jsonify({'success': False, 'message': 'Usuário ou senha inválidos'})
        except Exception as e:
            return jsonify({'success': False, 'message': 'Erro ao processar login'})
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ==================== FUNCIONÁRIOS ====================
@app.route('/funcionarios')
def funcionarios():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('funcionarios.html')

@app.route('/api/funcionarios', methods=['GET'])
def api_funcionarios_listar():
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'todos')
    
    query = Funcionario.query
    
    if busca:
        query = query.filter(
            (Funcionario.nome.ilike(f'%{busca}%')) |
            (Funcionario.cargo.ilike(f'%{busca}%'))
        )
    
    if ativo == 'ativos':
        query = query.filter_by(ativo=True)
    elif ativo == 'inativos':
        query = query.filter_by(ativo=False)
    
    funcionarios = query.order_by(Funcionario.nome).all()
    
    return jsonify([{
        'id': f.id,
        'nome': f.nome,
        'cargo': f.cargo,
        'telefone': f.telefone,
        'celular': f.celular,
        'email': f.email,
        'cpf': f.cpf,
        'endereco': f.endereco,
        'cidade': f.cidade,
        'uf': f.uf,
        'cep': f.cep,
        'data_admissao': f.data_admissao.strftime('%Y-%m-%d') if f.data_admissao else None,
        'ativo': f.ativo
    } for f in funcionarios])

@app.route('/api/funcionarios', methods=['POST'])
def api_funcionarios_criar():
    data = request.json
    
    if 'id' in data and data['id']:
        funcionario = Funcionario.query.get(data['id'])
        if funcionario:
            funcionario.nome = data['nome']
            funcionario.cargo = data.get('cargo')
            funcionario.telefone = data.get('telefone')
            funcionario.celular = data.get('celular')
            funcionario.email = data.get('email')
            funcionario.endereco = data.get('endereco')
            funcionario.cidade = data.get('cidade')
            funcionario.uf = data.get('uf')
            funcionario.cep = data.get('cep')
            funcionario.cpf = data.get('cpf')
            funcionario.data_admissao = datetime.strptime(data['data_admissao'], '%Y-%m-%d').date() if data.get('data_admissao') else None
            funcionario.ativo = data.get('ativo', True)
            db.session.commit()
            return jsonify({'success': True, 'id': funcionario.id})
    
    funcionario = Funcionario(
        nome=data['nome'],
        cargo=data.get('cargo'),
        telefone=data.get('telefone'),
        celular=data.get('celular'),
        email=data.get('email'),
        endereco=data.get('endereco'),
        cidade=data.get('cidade'),
        uf=data.get('uf'),
        cep=data.get('cep'),
        cpf=data.get('cpf'),
        data_admissao=datetime.strptime(data['data_admissao'], '%Y-%m-%d').date() if data.get('data_admissao') else None,
        ativo=data.get('ativo', True)
    )
    
    db.session.add(funcionario)
    db.session.commit()
    
    return jsonify({'success': True, 'id': funcionario.id})

@app.route('/api/funcionarios/<int:id>', methods=['PUT'])
def api_funcionarios_atualizar(id):
    funcionario = Funcionario.query.get_or_404(id)
    data = request.json
    
    funcionario.nome = data['nome']
    funcionario.cargo = data.get('cargo')
    funcionario.telefone = data.get('telefone')
    funcionario.celular = data.get('celular')
    funcionario.email = data.get('email')
    funcionario.endereco = data.get('endereco')
    funcionario.cidade = data.get('cidade')
    funcionario.uf = data.get('uf')
    funcionario.cep = data.get('cep')
    funcionario.cpf = data.get('cpf')
    funcionario.data_admissao = datetime.strptime(data['data_admissao'], '%Y-%m-%d').date() if data.get('data_admissao') else None
    funcionario.ativo = data.get('ativo', True)
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/funcionarios/<int:id>', methods=['DELETE'])
def api_funcionarios_deletar(id):
    funcionario = Funcionario.query.get_or_404(id)
    db.session.delete(funcionario)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/exportar-funcionarios-excel')
def exportar_funcionarios_excel():
    from flask import send_file
    
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'ativos')
    
    query = Funcionario.query
    
    if busca:
        query = query.filter(Funcionario.nome.like(f'%{busca}%'))
    
    if ativo == 'ativos':
        query = query.filter(Funcionario.ativo == True)
    elif ativo == 'inativos':
        query = query.filter(Funcionario.ativo == False)
    
    funcionarios = query.order_by(Funcionario.nome).all()
    
    headers = ['ID', 'Nome', 'Cargo', 'CPF', 'Email', 'Telefone', 'Endereço', 'Cidade', 'UF', 'CEP', 'Data Admissão', 'Status']
    dados = []
    
    for f in funcionarios:
        dados.append([
            f.id,
            f.nome,
            f.cargo or '',
            f.cpf or '',
            f.email or '',
            f.telefone or '',
            f.endereco or '',
            f.cidade or '',
            f.uf or '',
            f.cep or '',
            f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '',
            'Ativo' if f.ativo else 'Inativo'
        ])
    
    output = criar_excel_padrao('Funcionários', headers, dados, 'funcionarios.xlsx')
    
    return send_file(output, download_name='funcionarios.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/importar-funcionarios-excel', methods=['POST'])
def importar_funcionarios_excel():
    """Importar funcionários em massa a partir de planilha Excel"""
    try:
        import pandas as pd
        
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use apenas .xlsx ou .xls'})
        
        # Ler o arquivo Excel
        df = pd.read_excel(arquivo, header=None)
        
        # Procurar a linha que contém "Nome" ou "Cargo" (cabeçalhos reais)
        linha_cabecalho = 0
        for i in range(min(10, len(df))):
            linha_str = ' '.join([str(x).upper() for x in df.iloc[i].values if pd.notna(x)])
            if 'NOME' in linha_str or 'CARGO' in linha_str or 'CPF' in linha_str:
                linha_cabecalho = i
                break
        
        # Reler o arquivo com o cabeçalho correto
        df = pd.read_excel(arquivo, header=linha_cabecalho)
        
        # Remover colunas completamente vazias (Unnamed)
        df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed', na=False)]
        
        # Limpar nomes das colunas
        colunas_validas = []
        for col in df.columns:
            col_str = str(col).strip()
            if col_str and col_str.lower() not in ['nan', 'none', '']:
                colunas_validas.append(col)
        
        df = df[colunas_validas]
        df.columns = [str(col).strip() for col in df.columns]
        
        # Mapear colunas
        colunas_esperadas = {
            'nome': ['Nome', 'NOME', 'Funcionário', 'FUNCIONÁRIO'],
            'cargo': ['Cargo', 'CARGO'],
            'cpf': ['CPF', 'Cpf'],
            'email': ['Email', 'EMAIL', 'E-mail'],
            'telefone': ['Telefone', 'TELEFONE', 'Tel', 'Celular', 'CELULAR'],
            'endereco': ['Endereço', 'Endereco', 'ENDEREÇO', 'ENDERECO'],
            'cidade': ['Cidade', 'CIDADE'],
            'uf': ['UF', 'Estado', 'ESTADO'],
            'cep': ['CEP', 'Cep'],
            'data_admissao': ['Data Admissão', 'Data Admissao', 'Admissão', 'Admissao'],
            'status': ['Status', 'STATUS', 'Ativo', 'ATIVO']
        }
        
        mapeamento = {}
        for campo, variacoes in colunas_esperadas.items():
            for variacao in variacoes:
                coluna_encontrada = None
                for col in df.columns:
                    if str(col).strip().upper() == str(variacao).strip().upper():
                        coluna_encontrada = col
                        break
                if coluna_encontrada:
                    mapeamento[campo] = coluna_encontrada
                    break
        
        if 'nome' not in mapeamento:
            return jsonify({
                'success': False,
                'message': f'Coluna "Nome" não encontrada. Colunas: {", ".join(df.columns.tolist())}'
            })
        
        def pegar_valor(row, campo, tipo='str', padrao=None):
            try:
                if campo not in mapeamento:
                    return padrao
                valor = row[mapeamento[campo]]
                if pd.isna(valor) or valor == '':
                    return padrao
                if tipo == 'str':
                    return str(valor).strip() if valor else padrao
                elif tipo == 'date':
                    if isinstance(valor, str):
                        return datetime.strptime(valor, '%d/%m/%Y').date()
                    return valor
                return valor
            except:
                return padrao
        
        importados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                nome = pegar_valor(row, 'nome', 'str', '')
                if not nome:
                    continue
                
                funcionario = Funcionario(
                    nome=nome,
                    cargo=pegar_valor(row, 'cargo', 'str', None),
                    cpf=pegar_valor(row, 'cpf', 'str', None),
                    email=pegar_valor(row, 'email', 'str', None),
                    telefone=pegar_valor(row, 'telefone', 'str', None),
                    endereco=pegar_valor(row, 'endereco', 'str', None),
                    cidade=pegar_valor(row, 'cidade', 'str', None),
                    uf=pegar_valor(row, 'uf', 'str', None),
                    cep=pegar_valor(row, 'cep', 'str', None),
                    data_admissao=pegar_valor(row, 'data_admissao', 'date', None),
                    ativo=True
                )
                
                db.session.add(funcionario)
                importados += 1
            except Exception as e:
                erros.append(f"Linha {index + 2}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'importados': importados,
            'erros': erros if erros else None,
            'message': f'{importados} funcionários importados com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

# ==================== FORNECEDORES ====================

# ==================== FORNECEDORES ====================

@app.route('/fornecedores')
def fornecedores():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('fornecedores.html')

@app.route('/api/fornecedores', methods=['GET'])
def api_fornecedores_listar():
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'todos')
    filtro_categoria = request.args.get('filtro_categoria', 'false')
    
    query = Fornecedor.query
    
    if busca:
        # Se filtro_categoria está ativo, buscar APENAS pela categoria
        if filtro_categoria == 'true':
            query = query.filter(Fornecedor.categoria.ilike(f'%{busca}%'))
        else:
            # Busca normal por nome ou CNPJ
            query = query.filter(
                (Fornecedor.nome.ilike(f'%{busca}%')) |
                (Fornecedor.cnpj.ilike(f'%{busca}%'))
            )
    
    if ativo == 'ativos':
        query = query.filter_by(ativo=True)
    elif ativo == 'inativos':
        query = query.filter_by(ativo=False)
    
    fornecedores = query.order_by(Fornecedor.nome).all()
    
    return jsonify([{
        'codigo': f.codigo,
        'nome': f.nome,
        'telefone': f.telefone,
        'celular': f.celular,
        'email': f.email,
        'categoria': f.categoria,
        'endereco': f.endereco,
        'cidade': f.cidade,
        'uf': f.uf,
        'cep': f.cep,
        'cnpj': f.cnpj,
        'contato': f.contato,
        'ativo': f.ativo
    } for f in fornecedores])

@app.route('/api/fornecedores', methods=['POST'])
def api_fornecedores_criar():
    data = request.json
    
    fornecedor = Fornecedor(
        nome=data['nome'],
        telefone=data.get('telefone'),
        celular=data.get('celular'),
        email=data.get('email'),
        categoria=data.get('categoria'),
        endereco=data.get('endereco'),
        cidade=data.get('cidade'),
        uf=data.get('uf'),
        cep=data.get('cep'),
        cnpj=data.get('cnpj'),
        contato=data.get('contato'),
        ativo=data.get('ativo', True)
    )
    
    db.session.add(fornecedor)
    db.session.commit()
    
    return jsonify({'success': True, 'codigo': fornecedor.codigo})

@app.route('/api/fornecedores/<int:codigo>', methods=['PUT'])
def api_fornecedores_atualizar(codigo):
    fornecedor = Fornecedor.query.get_or_404(codigo)
    data = request.json
    
    fornecedor.nome = data['nome']
    fornecedor.telefone = data.get('telefone')
    fornecedor.celular = data.get('celular')
    fornecedor.email = data.get('email')
    fornecedor.categoria = data.get('categoria')
    fornecedor.endereco = data.get('endereco')
    fornecedor.cidade = data.get('cidade')
    fornecedor.uf = data.get('uf')
    fornecedor.cep = data.get('cep')
    fornecedor.cnpj = data.get('cnpj')
    fornecedor.contato = data.get('contato')
    fornecedor.ativo = data.get('ativo', True)
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/fornecedores/<int:codigo>', methods=['DELETE'])
def api_fornecedores_deletar(codigo):
    fornecedor = Fornecedor.query.get_or_404(codigo)
    db.session.delete(fornecedor)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/exportar-fornecedores-excel')
def exportar_fornecedores_excel():
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'ativos')
    
    query = Fornecedor.query
    
    if busca:
        query = query.filter(
            (Fornecedor.nome.ilike(f'%{busca}%')) |
            (Fornecedor.cnpj.ilike(f'%{busca}%'))
        )
    
    if ativo == 'ativos':
        query = query.filter(Fornecedor.ativo == True)
    elif ativo == 'inativos':
        query = query.filter(Fornecedor.ativo == False)
    
    fornecedores = query.order_by(Fornecedor.nome).all()
    
    headers = ['Código', 'Nome', 'CNPJ', 'Telefone', 'Celular', 'Email', 'Categoria', 'Endereço', 'Cidade', 'UF', 'CEP', 'Contato', 'Status']
    dados = []
    
    for f in fornecedores:
        dados.append([
            f.codigo,
            f.nome,
            f.cnpj or '',
            f.telefone or '',
            f.celular or '',
            f.email or '',
            f.categoria or '',
            f.endereco or '',
            f.cidade or '',
            f.uf or '',
            f.cep or '',
            f.contato or '',
            'Ativo' if f.ativo else 'Inativo'
        ])
    
    output = criar_excel_padrao('Fornecedores', headers, dados, 'fornecedores.xlsx')
    
    return send_file(output, download_name='fornecedores.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/importar-fornecedores-excel', methods=['POST'])
def importar_fornecedores_excel():
    """Importar fornecedores em massa a partir de planilha Excel"""
    try:
        import pandas as pd
        
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use apenas .xlsx ou .xls'})
        
        # Ler o arquivo Excel
        df = pd.read_excel(arquivo, header=None)
        
        # Procurar a linha que contém "Nome" ou "CNPJ" (cabeçalhos reais)
        linha_cabecalho = 0
        for i in range(min(10, len(df))):
            linha_str = ' '.join([str(x).upper() for x in df.iloc[i].values if pd.notna(x)])
            if 'NOME' in linha_str or 'CNPJ' in linha_str or 'CÓDIGO' in linha_str:
                linha_cabecalho = i
                break
        
        # Reler o arquivo com o cabeçalho correto
        df = pd.read_excel(arquivo, header=linha_cabecalho)
        
        # Remover colunas completamente vazias (Unnamed)
        df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed', na=False)]
        
        # Limpar nomes das colunas
        colunas_validas = []
        for col in df.columns:
            col_str = str(col).strip()
            if col_str and col_str.lower() not in ['nan', 'none', '']:
                colunas_validas.append(col)
        
        df = df[colunas_validas]
        df.columns = [str(col).strip() for col in df.columns]
        
        # Mapear colunas
        colunas_esperadas = {
            'nome': ['Nome', 'NOME', 'Fornecedor', 'FORNECEDOR', 'Razão Social'],
            'cnpj': ['CNPJ', 'Cnpj'],
            'telefone': ['Telefone', 'TELEFONE', 'Tel'],
            'celular': ['Celular', 'CELULAR', 'Cel'],
            'email': ['Email', 'EMAIL', 'E-mail'],
            'categoria': ['Categoria', 'CATEGORIA', 'Oficina', 'OFICINA', 'oficina', 'OFICINA'],
            'endereco': ['Endereço', 'Endereco', 'ENDEREÇO', 'ENDERECO'],
            'cidade': ['Cidade', 'CIDADE'],
            'uf': ['UF', 'Estado', 'ESTADO'],
            'cep': ['CEP', 'Cep'],
            'contato': ['Contato', 'CONTATO', 'Pessoa Contato']
        }
        
        mapeamento = {}
        for campo, variacoes in colunas_esperadas.items():
            for variacao in variacoes:
                coluna_encontrada = None
                for col in df.columns:
                    if str(col).strip().upper() == str(variacao).strip().upper():
                        coluna_encontrada = col
                        break
                if coluna_encontrada:
                    mapeamento[campo] = coluna_encontrada
                    break
        
        if 'nome' not in mapeamento:
            return jsonify({
                'success': False,
                'message': f'Coluna "Nome" não encontrada. Colunas: {", ".join(df.columns.tolist())}'
            })
        
        def pegar_valor(row, campo, tipo='str', padrao=None):
            try:
                if campo not in mapeamento:
                    return padrao
                valor = row[mapeamento[campo]]
                if pd.isna(valor) or valor == '':
                    return padrao
                if tipo == 'str':
                    return str(valor).strip() if valor else padrao
                return valor
            except:
                return padrao
        
        importados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                nome = pegar_valor(row, 'nome', 'str', '')
                if not nome:
                    continue
                
                fornecedor = Fornecedor(
                    nome=nome,
                    cnpj=pegar_valor(row, 'cnpj', 'str', None),
                    telefone=pegar_valor(row, 'telefone', 'str', None),
                    celular=pegar_valor(row, 'celular', 'str', None),
                    email=pegar_valor(row, 'email', 'str', None),
                    categoria=pegar_valor(row, 'categoria', 'str', None),
                    endereco=pegar_valor(row, 'endereco', 'str', None),
                    cidade=pegar_valor(row, 'cidade', 'str', None),
                    uf=pegar_valor(row, 'uf', 'str', None),
                    cep=pegar_valor(row, 'cep', 'str', None),
                    contato=pegar_valor(row, 'contato', 'str', None),
                    ativo=True
                )
                
                db.session.add(fornecedor)
                importados += 1
            except Exception as e:
                erros.append(f"Linha {index + 2}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'importados': importados,
            'erros': erros if erros else None,
            'message': f'{importados} fornecedores importados com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

# ==================== CLIENTES ====================

@app.route('/clientes')
def clientes():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('clientes.html')

@app.route('/api/clientes', methods=['GET'])
def api_clientes_listar():
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'todos')
    
    query = Cliente.query
    
    if busca:
        query = query.filter(
            (Cliente.nome.ilike(f'%{busca}%')) |
            (Cliente.cpfcnpj.ilike(f'%{busca}%'))
        )
    
    if ativo == 'ativos':
        query = query.filter_by(ativo=True)
    elif ativo == 'inativos':
        query = query.filter_by(ativo=False)
    
    clientes = query.order_by(Cliente.nome).all()
    
    return jsonify([{
        'id': c.id,
        'nome': c.nome,
        'email': c.email,
        'cpfcnpj': c.cpfcnpj,
        'telefone': c.telefone,
        'celular': c.celular,
        'endereco': c.endereco,
        'cidade': c.cidade,
        'uf': c.uf,
        'cep': c.cep,
        'anotacoes': c.anotacoes,
        'ativo': c.ativo
    } for c in clientes])

@app.route('/api/clientes', methods=['POST'])
def api_clientes_criar():
    data = request.json
    
    # Validar CPF/CNPJ duplicado
    cpfcnpj = data.get('cpfcnpj', '').strip()
    if cpfcnpj:
        cliente_existente = Cliente.query.filter_by(cpfcnpj=cpfcnpj).first()
        if cliente_existente:
            return jsonify({'success': False, 'message': 'CPF/CNPJ já cadastrado!'}), 400
    
    cliente = Cliente(
        nome=data['nome'],
        email=data.get('email'),
        cpfcnpj=cpfcnpj,
        telefone=data.get('telefone'),
        celular=data.get('celular'),
        endereco=data.get('endereco'),
        cidade=data.get('cidade'),
        uf=data.get('uf'),
        cep=data.get('cep'),
        anotacoes=data.get('anotacoes'),
        ativo=data.get('ativo', True)
    )
    
    db.session.add(cliente)
    db.session.commit()
    
    return jsonify({'success': True, 'id': cliente.id})

@app.route('/api/clientes/<int:id>', methods=['PUT'])
def api_clientes_atualizar(id):
    cliente = Cliente.query.get_or_404(id)
    data = request.json
    
    # Validar CPF/CNPJ duplicado (exceto o próprio cliente)
    cpfcnpj = data.get('cpfcnpj', '').strip()
    if cpfcnpj:
        cliente_existente = Cliente.query.filter(Cliente.cpfcnpj == cpfcnpj, Cliente.id != id).first()
        if cliente_existente:
            return jsonify({'success': False, 'message': 'CPF/CNPJ já cadastrado!'}), 400
    
    cliente.nome = data['nome']
    cliente.email = data.get('email')
    cliente.cpfcnpj = cpfcnpj
    cliente.telefone = data.get('telefone')
    cliente.celular = data.get('celular')
    cliente.endereco = data.get('endereco')
    cliente.cidade = data.get('cidade')
    cliente.uf = data.get('uf')
    cliente.cep = data.get('cep')
    cliente.anotacoes = data.get('anotacoes')
    cliente.ativo = data.get('ativo', True)
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/clientes/<int:id>', methods=['DELETE'])
def api_clientes_deletar(id):
    cliente = Cliente.query.get_or_404(id)
    db.session.delete(cliente)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/exportar-clientes-excel')
def exportar_clientes_excel():
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'ativos')
    
    query = Cliente.query
    
    if busca:
        query = query.filter(
            (Cliente.nome.ilike(f'%{busca}%')) |
            (Cliente.cpfcnpj.ilike(f'%{busca}%'))
        )
    
    if ativo == 'ativos':
        query = query.filter(Cliente.ativo == True)
    elif ativo == 'inativos':
        query = query.filter(Cliente.ativo == False)
    
    clientes = query.order_by(Cliente.nome).all()
    
    headers = ['ID', 'Nome', 'Email', 'CPF/CNPJ', 'Telefone', 'Celular', 'Endereço', 'Cidade', 'UF', 'CEP', 'Anotações', 'Status']
    dados = []
    
    for c in clientes:
        dados.append([
            c.id,
            c.nome,
            c.email or '',
            c.cpfcnpj or '',
            c.telefone or '',
            c.celular or '',
            c.endereco or '',
            c.cidade or '',
            c.uf or '',
            c.cep or '',
            c.anotacoes or '',
            'Ativo' if c.ativo else 'Inativo'
        ])
    
    output = criar_excel_padrao('Clientes', headers, dados, 'clientes.xlsx')
    
    return send_file(output, download_name='clientes.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/importar-clientes-excel', methods=['POST'])
def importar_clientes_excel():
    try:
        import pandas as pd
        
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use apenas .xlsx ou .xls'})
        
        df = pd.read_excel(arquivo, header=None)
        
        linha_cabecalho = 0
        for i in range(min(10, len(df))):
            linha_str = ' '.join([str(x).upper() for x in df.iloc[i].values if pd.notna(x)])
            if 'NOME' in linha_str or 'ID' in linha_str or 'EMAIL' in linha_str:
                linha_cabecalho = i
                break
        
        df = pd.read_excel(arquivo, header=linha_cabecalho)
        df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed', na=False)]
        
        colunas_validas = []
        for col in df.columns:
            col_str = str(col).strip()
            if col_str and col_str.lower() not in ['nan', 'none', '']:
                colunas_validas.append(col)
        
        df = df[colunas_validas]
        df.columns = [str(col).strip() for col in df.columns]
        
        colunas_esperadas = {
            'nome': ['Nome', 'NOME', 'Cliente', 'CLIENTE'],
            'email': ['Email', 'EMAIL', 'E-mail'],
            'cpfcnpj': ['CPF/CNPJ', 'CPF', 'CNPJ', 'Cpf', 'Cnpj', 'Documento'],
            'telefone': ['Telefone', 'TELEFONE', 'Fone Comercial', 'Tel'],
            'celular': ['Celular', 'CELULAR', 'Fone Celular', 'Cel'],
            'endereco': ['Endereço', 'Endereco', 'ENDEREÇO', 'ENDERECO'],
            'cidade': ['Cidade', 'CIDADE'],
            'uf': ['UF', 'Estado', 'ESTADO'],
            'cep': ['CEP', 'Cep'],
            'anotacoes': ['Anotações', 'Anotacoes', 'ANOTAÇÕES', 'ANOTACOES', 'Observações', 'Observacoes']
        }
        
        mapeamento = {}
        for campo, variacoes in colunas_esperadas.items():
            for variacao in variacoes:
                coluna_encontrada = None
                for col in df.columns:
                    if str(col).strip().upper() == str(variacao).strip().upper():
                        coluna_encontrada = col
                        break
                if coluna_encontrada:
                    mapeamento[campo] = coluna_encontrada
                    break
        
        if 'nome' not in mapeamento:
            return jsonify({
                'success': False,
                'message': f'Coluna "Nome" não encontrada. Colunas: {", ".join(df.columns.tolist())}'
            })
        
        def pegar_valor(row, campo, tipo='str', padrao=None):
            try:
                if campo not in mapeamento:
                    return padrao
                valor = row[mapeamento[campo]]
                if pd.isna(valor) or valor == '':
                    return padrao
                if tipo == 'str':
                    return str(valor).strip() if valor else padrao
                return valor
            except:
                return padrao
        
        importados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                nome = pegar_valor(row, 'nome', 'str', '')
                if not nome:
                    continue
                
                cpfcnpj = pegar_valor(row, 'cpfcnpj', 'str', None)
                
                # Validar CPF/CNPJ duplicado
                if cpfcnpj:
                    cliente_existente = Cliente.query.filter_by(cpfcnpj=cpfcnpj).first()
                    if cliente_existente:
                        erros.append(f"Linha {index + 2}: CPF/CNPJ {cpfcnpj} já existe")
                        continue
                
                cliente = Cliente(
                    nome=nome,
                    email=pegar_valor(row, 'email', 'str', None),
                    cpfcnpj=cpfcnpj,
                    telefone=pegar_valor(row, 'telefone', 'str', None),
                    celular=pegar_valor(row, 'celular', 'str', None),
                    endereco=pegar_valor(row, 'endereco', 'str', None),
                    cidade=pegar_valor(row, 'cidade', 'str', None),
                    uf=pegar_valor(row, 'uf', 'str', None),
                    cep=pegar_valor(row, 'cep', 'str', None),
                    anotacoes=pegar_valor(row, 'anotacoes', 'str', None),
                    ativo=True
                )
                
                db.session.add(cliente)
                importados += 1
            except Exception as e:
                erros.append(f"Linha {index + 2}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'importados': importados,
            'erros': erros if erros else None,
            'message': f'{importados} clientes importados com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

# ==================== PRODUTOS ====================

@app.route('/api/upload-foto', methods=['POST'])
def upload_foto():
    if 'foto' not in request.files:
        return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
    
    file = request.files['foto']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
    
    if file:
        # Gerar nome seguro para o arquivo
        filename = secure_filename(file.filename)
        
        # Salvar o arquivo
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        return jsonify({'success': True, 'filename': filename})
    
    return jsonify({'success': False, 'message': 'Erro ao processar arquivo'})

@app.route('/api/foto/<filename>')
def get_foto(filename):
    """Servir fotos de produtos"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(filename))
        if os.path.exists(filepath):
            return send_file(filepath)
        else:
            # Se não encontrar, retorna a imagem padrão
            default_path = os.path.join(app.config['UPLOAD_FOLDER'], 'imagem-padrao.jpg')
            if os.path.exists(default_path):
                return send_file(default_path)
            return jsonify({'error': 'Arquivo não encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/produtos')
def produtos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('produtos.html')

@app.route('/api/produtos', methods=['GET'])
def api_produtos_listar():
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'todos')
    
    query = Produto.query
    
    if busca:
        # Normalizar busca removendo acentos
        busca_normalizada = remover_acentos(busca).lower()
        
        # Buscar em todos os produtos
        todos_produtos = query.all()
        produtos_filtrados = []
        
        for p in todos_produtos:
            # Normalizar campos do produto
            descricao_norm = remover_acentos(p.descricao or '').lower()
            codigo_norm = str(p.codigo or '').lower()
            codigo_barras_norm = remover_acentos(p.codigo_barras or '').lower()
            
            # Verificar se a busca está em algum campo
            if (busca_normalizada in descricao_norm or 
                busca_normalizada in codigo_norm or 
                busca_normalizada in codigo_barras_norm):
                produtos_filtrados.append(p)
        
        produtos = produtos_filtrados
    else:
        produtos = query.all()
    
    # Aplicar filtro de ativo
    if ativo == 'ativos':
        produtos = [p for p in produtos if p.ativo]
    elif ativo == 'inativos':
        produtos = [p for p in produtos if not p.ativo]
    
    # Ordenar por descrição
    produtos = sorted(produtos, key=lambda p: p.descricao or '')
    
    # Limitar a 100 produtos
    produtos = produtos[:100]
    
    # Buscar todos os empréstimos de uma vez em vez de fazer query por produto
    emprestimos_ativos = Emprestimo.query.filter_by(status='EMPRESTADO').all()
    emprestimos_por_produto = {}
    for e in emprestimos_ativos:
        if e.cod_produto not in emprestimos_por_produto:
            emprestimos_por_produto[e.cod_produto] = 0
        emprestimos_por_produto[e.cod_produto] += (e.quantidade or 1)
    
    resultado = []
    for p in produtos:
        resultado.append({
            'codigo': p.codigo,
            'descricao': p.descricao,
            'codigo_barras': p.codigo_barras,
            'unidade': p.unidade,
            'marca': p.marca,
            'categoria': p.categoria,
            'preco_compra': p.preco_compra,
            'preco_venda': p.preco_venda,
            'estoque_minimo': p.estoque_minimo,
            'estoque_maximo': p.estoque_maximo,
            'estoque_atual': p.estoque_atual,
            'foto': p.foto,
            'fornecedor': p.fornecedor,
            'observacao': p.observacao,
            'ativo': p.ativo,
            'emprestados': emprestimos_por_produto.get(p.codigo, 0)
        })
    
    return jsonify(resultado)

@app.route('/api/produtos', methods=['POST'])
def api_produtos_criar():
    data = request.json
    
    # Se vier com codigo, é alteração (não deveria usar POST, mas vamos aceitar)
    if 'codigo' in data and data['codigo']:
        produto = Produto.query.get(data['codigo'])
        if produto:
            # Atualizar produto existente
            produto.descricao = data['descricao']
            produto.codigo_barras = data.get('codigo_barras')
            produto.unidade = data.get('unidade')
            produto.marca = data.get('marca')
            produto.categoria = data.get('categoria')
            produto.preco_compra = data.get('preco_compra', 0)
            produto.preco_venda = data.get('preco_venda', 0)
            produto.estoque_minimo = data.get('estoque_minimo', 0)
            produto.estoque_maximo = data.get('estoque_maximo', 0)
            produto.fornecedor = data.get('fornecedor')
            produto.observacao = data.get('observacao')
            produto.foto = data.get('foto')
            produto.ativo = data.get('ativo', True)
            
            db.session.commit()
            return jsonify({'success': True, 'codigo': produto.codigo})
    
    # Criar novo produto (ID automático)
    produto = Produto(
        descricao=data['descricao'],
        codigo_barras=data.get('codigo_barras'),
        unidade=data.get('unidade'),
        marca=data.get('marca'),
        categoria=data.get('categoria'),
        preco_compra=data.get('preco_compra', 0),
        preco_venda=data.get('preco_venda', 0),
        estoque_minimo=data.get('estoque_minimo', 0),
        estoque_maximo=data.get('estoque_maximo', 0),
        estoque_atual=data.get('estoque_atual', 0),
        fornecedor=data.get('fornecedor'),
        observacao=data.get('observacao'),
        foto=data.get('foto'),
        ativo=data.get('ativo', True)
    )
    
    db.session.add(produto)
    db.session.commit()
    
    return jsonify({'success': True, 'codigo': produto.codigo})

@app.route('/api/produtos/<int:codigo>', methods=['PUT'])
def api_produtos_atualizar(codigo):
    produto = Produto.query.get_or_404(codigo)
    data = request.json
    
    produto.descricao = data['descricao']
    produto.unidade = data.get('unidade')
    produto.marca = data.get('marca')
    produto.categoria = data.get('categoria')
    produto.preco_compra = data.get('preco_compra', 0)
    produto.preco_venda = data.get('preco_venda', 0)
    produto.estoque_minimo = data.get('estoque_minimo', 0)
    produto.estoque_maximo = data.get('estoque_maximo', 0)
    produto.fornecedor = data.get('fornecedor')
    produto.foto = data.get('foto')
    produto.ativo = data.get('ativo', True)
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/produtos/<int:codigo>', methods=['GET'])
def api_produtos_obter(codigo):
    """Obter dados de um produto específico"""
    produto = Produto.query.get_or_404(codigo)
    
    return jsonify({
        'codigo': produto.codigo,
        'descricao': produto.descricao,
        'codigo_barras': produto.codigo_barras,
        'unidade': produto.unidade,
        'marca': produto.marca,
        'categoria': produto.categoria,
        'fornecedor': produto.fornecedor,
        'preco_compra': produto.preco_compra,
        'preco_venda': produto.preco_venda,
        'estoque_minimo': produto.estoque_minimo,
        'estoque_maximo': produto.estoque_maximo,
        'estoque_atual': produto.estoque_atual,
        'observacao': produto.observacao,
        'foto': produto.foto,
        'ativo': produto.ativo
    })

@app.route('/api/produtos/<int:codigo>', methods=['DELETE'])
def api_produtos_deletar(codigo):
    produto = Produto.query.get_or_404(codigo)
    db.session.delete(produto)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/produtos/<int:codigo>/estoque-disponivel', methods=['GET'])
def api_produtos_estoque_disponivel(codigo):
    produto = Produto.query.get_or_404(codigo)
    
    # Somar quantidades emprestadas
    emprestimos_ativos = Emprestimo.query.filter_by(
        cod_produto=codigo,
        status='EMPRESTADO'
    ).all()
    emprestados = sum(e.quantidade or 1 for e in emprestimos_ativos)
    
    # Calcular disponível
    disponivel = produto.estoque_atual - emprestados
    
    return jsonify({
        'total': produto.estoque_atual,
        'emprestado': emprestados,
        'disponivel': disponivel
    })

# ==================== MOVIMENTAÇÕES ====================

class Movimentacao(db.Model):
    __tablename__ = 'movimentacoes'
    
    id = db.Column(db.Integer, primary_key=True)
    produto_codigo = db.Column(db.Integer, db.ForeignKey('produtos.codigo'), nullable=False, index=True)
    tipo = db.Column(db.String(20), nullable=False, index=True)  # ENTRADA, SAIDA, EMPRESTIMO, DEVOLUCAO
    quantidade = db.Column(db.Integer, nullable=False)
    valor_unitario = db.Column(db.Float, default=0)
    valor_total = db.Column(db.Float, default=0)
    nota_fiscal = db.Column(db.String(100), index=True)
    operacao = db.Column(db.String(100))
    observacao = db.Column(db.Text)
    solicitante_tipo = db.Column(db.String(50))  # FUNCIONARIO, CLIENTE, FORNECEDOR
    solicitante_nome = db.Column(db.String(200))  # Nome do solicitante
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), index=True)
    data_movimentacao = db.Column(db.DateTime, default=datetime.now, index=True)
    lote_id = db.Column(db.String(50), index=True)  # Agrupa entradas múltiplas da empresa MASTER

@app.route('/api/movimentacoes/entrada', methods=['POST'])
def api_entrada():
    data = request.json
    produto_codigo = data['produto_codigo']
    quantidade = int(data['quantidade'])
    
    # Buscar produto
    produto = Produto.query.get_or_404(produto_codigo)
    
    # Criar movimentação
    movimentacao = Movimentacao(
        produto_codigo=produto_codigo,
        tipo='ENTRADA',
        quantidade=quantidade,
        valor_unitario=data.get('valor_unitario', 0),
        valor_total=data.get('valor_total', 0),
        nota_fiscal=data.get('nota_fiscal'),
        operacao=data.get('operacao'),
        observacao=data.get('observacao'),
        solicitante_tipo=data.get('solicitante_tipo', 'FORNECEDOR'),
        solicitante_nome=data.get('solicitante_nome', ''),
        usuario_id=session.get('usuario_id')
    )
    
    # Atualizar estoque
    produto.estoque_atual += quantidade
    
    db.session.add(movimentacao)
    db.session.commit()
    
    return jsonify({'success': True, 'estoque_atual': produto.estoque_atual})

@app.route('/api/movimentacoes/saida', methods=['POST'])
def api_saida():
    data = request.json
    produto_codigo = data['produto_codigo']
    quantidade = int(data['quantidade'])
    
    # Buscar produto
    produto = Produto.query.get_or_404(produto_codigo)
    
    # Verificar se há estoque suficiente
    if produto.estoque_atual < quantidade:
        return jsonify({'success': False, 'message': 'Estoque insuficiente!'}), 400
    
    # Criar movimentação
    movimentacao = Movimentacao(
        produto_codigo=produto_codigo,
        tipo='SAIDA',
        quantidade=quantidade,
        valor_unitario=data.get('valor_unitario', 0),
        valor_total=data.get('valor_total', 0),
        nota_fiscal=data.get('nota_fiscal'),
        operacao=data.get('operacao'),
        observacao=data.get('observacao'),
        solicitante_tipo=data.get('solicitante_tipo'),  # FUNCIONARIO, CLIENTE, FORNECEDOR
        solicitante_nome=data.get('solicitante_nome'),  # Nome do solicitante
        usuario_id=session.get('usuario_id')
    )
    
    # Atualizar estoque
    produto.estoque_atual -= quantidade
    
    db.session.add(movimentacao)
    db.session.commit()
    
    return jsonify({'success': True, 'estoque_atual': produto.estoque_atual})

@app.route('/api/movimentacoes/<int:produto_codigo>', methods=['GET'])
def api_movimentacoes_produto(produto_codigo):
    # Limitar a 100 movimentações mais recentes para evitar carregar tudo
    movimentacoes = Movimentacao.query.filter_by(produto_codigo=produto_codigo).order_by(Movimentacao.data_movimentacao.desc()).limit(100).all()
    
    return jsonify([{
        'id': m.id,
        'tipo': m.tipo,
        'quantidade': m.quantidade,
        'valor_unitario': m.valor_unitario,
        'valor_total': m.valor_total,
        'nota_fiscal': m.nota_fiscal,
        'operacao': m.operacao,
        'observacao': m.observacao,
        'data_movimentacao': m.data_movimentacao.strftime('%d/%m/%Y %H:%M')
    } for m in movimentacoes])

@app.route('/api/movimentacoes', methods=['GET'])
def api_movimentacoes_listar():
    # Filtros
    data_inicial = request.args.get('data_inicial')
    data_final = request.args.get('data_final')
    tipo = request.args.get('tipo')
    mes = request.args.get('mes')
    nota_fiscal = request.args.get('nota_fiscal')
    
    # Query base
    query = Movimentacao.query
    
    # Aplicar filtros
    if data_inicial:
        query = query.filter(Movimentacao.data_movimentacao >= data_inicial)
    if data_final:
        query = query.filter(Movimentacao.data_movimentacao <= data_final)
    if tipo and tipo != 'TODOS':
        query = query.filter(Movimentacao.tipo == tipo)
    if mes and mes != 'TODOS':
        from sqlalchemy import extract
        query = query.filter(extract('month', Movimentacao.data_movimentacao) == int(mes))
    if nota_fiscal:
        query = query.filter(Movimentacao.nota_fiscal.like(f'%{nota_fiscal}%'))
    
    # Ordenar por data decrescente
    movimentacoes = query.order_by(Movimentacao.data_movimentacao.desc()).all()
    
    # Retornar todas as movimentações individualmente (o frontend agrupa por nota fiscal)
    resultado = []
    for m in movimentacoes:
        produto = Produto.query.get(m.produto_codigo)
        
        # Buscar usuário real
        usuario_nome = 'Sistema'
        if m.usuario_id:
            usuario = Usuario.query.get(m.usuario_id)
            if usuario:
                usuario_nome = usuario.nome
        
        # Buscar solicitante/operação
        operacao_pessoa = ''
        if m.tipo == 'ENTRADA':
            if hasattr(m, 'solicitante_nome') and m.solicitante_nome:
                operacao_pessoa = m.solicitante_nome
            elif produto and produto.fornecedor:
                operacao_pessoa = produto.fornecedor
        else:
            if hasattr(m, 'solicitante_nome') and m.solicitante_nome:
                operacao_pessoa = m.solicitante_nome
        
        resultado.append({
            'id': m.id,
            'produto_codigo': m.produto_codigo,
            'produto_descricao': produto.descricao if produto else '',
            'tipo': m.tipo,
            'quantidade': m.quantidade,
            'valor_unitario': m.valor_unitario,
            'valor_total': m.valor_total,
            'nota_fiscal': m.nota_fiscal,
            'operacao': m.operacao,
            'observacao': m.observacao,
            'solicitante': operacao_pessoa,
            'solicitante_tipo': m.solicitante_tipo if hasattr(m, 'solicitante_tipo') else None,
            'usuario': usuario_nome,
            'data_movimentacao': m.data_movimentacao.strftime('%d/%m/%Y %H:%M') if m.data_movimentacao else ''
        })
    
    return jsonify(resultado)

@app.route('/api/entrada-multipla', methods=['POST'])
def api_entrada_multipla():
    data = request.json
    
    try:
        fornecedor_nome = data.get('fornecedor', '')
        itens = data.get('itens', [])
        
        if not itens:
            return jsonify({'success': False, 'message': 'Nenhum item para processar'}), 400
        
        # Verificar se é empresa MASTER (tem asterisco no nome)
        is_master = '*' in fornecedor_nome
        
        # Gerar lote_id único para empresa MASTER
        lote_id = None
        if is_master:
            from datetime import datetime
            lote_id = f"LOTE_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Buscar todos os produtos de uma vez (evita N queries)
        codigos = [int(item['codigo']) for item in itens]
        produtos_dict = {p.codigo: p for p in Produto.query.filter(Produto.codigo.in_(codigos)).all()}
        
        # Preparar movimentações e atualizações
        movimentacoes = []
        for item in itens:
            produto_codigo = int(item['codigo'])
            quantidade = int(item['quantidade'])
            
            if produto_codigo not in produtos_dict:
                return jsonify({'success': False, 'message': f'Produto {produto_codigo} não encontrado'}), 404
            
            produto = produtos_dict[produto_codigo]
            
            # Criar movimentação
            movimentacao = Movimentacao(
                produto_codigo=produto_codigo,
                tipo='ENTRADA',
                quantidade=quantidade,
                valor_unitario=float(item['preco']),
                valor_total=float(item['total']),
                nota_fiscal=data.get('nota_fiscal'),
                operacao=data.get('operacao'),
                observacao=data.get('observacao'),
                solicitante_tipo='FORNECEDOR',
                solicitante_nome=fornecedor_nome,
                usuario_id=session.get('usuario_id'),
                lote_id=lote_id  # Atribuir lote_id para empresa MASTER
            )
            
            movimentacoes.append(movimentacao)
            
            # Atualizar estoque do produto
            produto.estoque_atual += quantidade
        
        # Adicionar todas as movimentações de uma vez
        db.session.add_all(movimentacoes)
        
        # Commit único para tudo
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/movimentacoes/<int:id>', methods=['DELETE'])
def api_movimentacoes_deletar(id):
    try:
        movimentacao = Movimentacao.query.get_or_404(id)
        produto = Produto.query.get(movimentacao.produto_codigo)
        
        if produto:
            # Reverter o estoque
            if movimentacao.tipo == 'ENTRADA':
                produto.estoque_atual -= movimentacao.quantidade
            elif movimentacao.tipo == 'SAIDA':
                produto.estoque_atual += movimentacao.quantidade
        
        db.session.delete(movimentacao)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ==================== ESTOQUE ====================

@app.route('/estoque')
def estoque():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('estoque.html')

@app.route('/movimentacoes')
def movimentacoes():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('movimentacoes.html')

@app.route('/entrada-multipla')
def entrada_multipla():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('entrada_multipla.html')

@app.route('/testar_foto.html')
def testar_foto():
    return render_template('testar_foto.html')

@app.route('/teste_foto_simples')
def teste_foto_simples():
    return render_template('teste_foto_simples.html')

@app.route('/teste_foto_debug')
def teste_foto_debug():
    return render_template('teste_foto_debug.html')

# ==================== CADASTROS DIVERSOS ====================

@app.route('/cadastros-diversos')
def cadastros_diversos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('cadastros_diversos.html')

# ==================== EMPRÉSTIMOS ====================

@app.route('/emprestimos')
def emprestimos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('emprestimos.html')

@app.route('/emprestimo-multiplo')
def emprestimo_multiplo():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('emprestimo_multiplo.html')

@app.route('/api/emprestimos', methods=['GET'])
def api_emprestimos_listar():
    # Filtros
    data_inicial = request.args.get('data_inicial', '')
    data_final = request.args.get('data_final', '')
    descricao = request.args.get('descricao', '')
    status = request.args.get('status', '')
    funcionario = request.args.get('funcionario', '')
    
    emprestimos = Emprestimo.query.order_by(Emprestimo.data.desc()).all()
    
    # Filtrar por data
    if data_inicial or data_final:
        emprestimos_filtrados = []
        for emp in emprestimos:
            try:
                # Extrair apenas a data (DD/MM/YYYY) ignorando a hora
                data_str = emp.data.split(' ')[0] if ' ' in emp.data else emp.data
                data_emp = datetime.strptime(data_str, '%d/%m/%Y')
                
                # Se tem data_inicial E data_final: busca por intervalo
                if data_inicial and data_final:
                    data_ini = datetime.strptime(data_inicial, '%Y-%m-%d')
                    data_fin = datetime.strptime(data_final, '%Y-%m-%d')
                    if data_ini <= data_emp <= data_fin:
                        emprestimos_filtrados.append(emp)
                # Se tem APENAS data_inicial: busca EXATA
                elif data_inicial and not data_final:
                    data_ini = datetime.strptime(data_inicial, '%Y-%m-%d')
                    if data_emp.date() == data_ini.date():
                        emprestimos_filtrados.append(emp)
                # Se tem APENAS data_final: busca até essa data
                elif data_final and not data_inicial:
                    data_fin = datetime.strptime(data_final, '%Y-%m-%d')
                    if data_emp <= data_fin:
                        emprestimos_filtrados.append(emp)
            except:
                # Se der erro ao converter, inclui o empréstimo
                emprestimos_filtrados.append(emp)
        
        emprestimos = emprestimos_filtrados
    
    # Filtrar por descrição (sem acentos)
    if descricao:
        descricao_normalizada = remover_acentos(descricao.lower())
        emprestimos = [e for e in emprestimos if descricao_normalizada in remover_acentos((e.descricao_item or '').lower())]
    
    # Filtrar por status
    if status:
        emprestimos = [e for e in emprestimos if e.status == status]
    
    # Filtrar por funcionário (sem acentos)
    if funcionario:
        funcionario_normalizado = remover_acentos(funcionario.lower())
        emprestimos = [e for e in emprestimos if funcionario_normalizado in remover_acentos((e.funcionario or '').lower())]
    
    # AGRUPAR empréstimos por: data + funcionário + item + status
    # Isso faz com que múltiplos empréstimos do mesmo item na mesma data apareçam em UMA linha
    grupos = {}
    for emp in emprestimos:
        # Chave de agrupamento: data + funcionário + item + status
        chave = f"{emp.data}|{emp.funcionario}|{emp.descricao_item}|{emp.status}"
        
        if chave not in grupos:
            grupos[chave] = {
                'emprestimos': [],
                'quantidade_total': 0
            }
        
        grupos[chave]['emprestimos'].append(emp)
        grupos[chave]['quantidade_total'] += (emp.quantidade or 1)
    
    resultado = []
    hoje = datetime.now()
    hoje_date = hoje.date()
    
    for chave, grupo in grupos.items():
        # Usar o primeiro empréstimo do grupo como referência
        emp = grupo['emprestimos'][0]
        
        # Calcular dias
        dias = '---'
        if emp.status == 'EMPRESTADO':
            try:
                # Extrair apenas a data (DD/MM/YYYY) ignorando a hora
                data_str = emp.data.split(' ')[0] if ' ' in emp.data else emp.data
                data_emp = datetime.strptime(data_str, '%d/%m/%Y')
                dias = (datetime.now() - data_emp).days
            except:
                pass
        elif emp.data_devolucao:
            try:
                data_str = emp.data.split(' ')[0] if ' ' in emp.data else emp.data
                data_dev_str = emp.data_devolucao.split(' ')[0] if ' ' in emp.data_devolucao else emp.data_devolucao
                data_emp = datetime.strptime(data_str, '%d/%m/%Y')
                data_dev = datetime.strptime(data_dev_str, '%d/%m/%Y')
                dias = (data_dev - data_emp).days
            except:
                pass
        
        # Converter data para formato ISO para o input datetime-local
        data_iso = ''
        try:
            # Se tem hora: "21/01/2026 14:30"
            if ' ' in emp.data:
                data_iso = datetime.strptime(emp.data, '%d/%m/%Y %H:%M').strftime('%Y-%m-%dT%H:%M')
            else:
                # Se não tem hora: "21/01/2026"
                data_iso = datetime.strptime(emp.data, '%d/%m/%Y').strftime('%Y-%m-%dT00:00')
        except:
            pass
        
        # Converter data_devolucao para formato ISO
        data_devolucao_iso = ''
        if emp.data_devolucao:
            try:
                if ' ' in emp.data_devolucao:
                    data_devolucao_iso = datetime.strptime(emp.data_devolucao, '%d/%m/%Y %H:%M').strftime('%Y-%m-%dT%H:%M')
                else:
                    data_devolucao_iso = datetime.strptime(emp.data_devolucao, '%d/%m/%Y').strftime('%Y-%m-%dT00:00')
            except:
                pass
        
        # Converter data_devolucao_prevista para formato ISO
        data_devolucao_prevista_iso = ''
        if emp.data_devolucao_prevista:
            try:
                if ' ' in emp.data_devolucao_prevista:
                    data_devolucao_prevista_iso = datetime.strptime(emp.data_devolucao_prevista, '%d/%m/%Y %H:%M').strftime('%Y-%m-%dT%H:%M')
                else:
                    data_devolucao_prevista_iso = datetime.strptime(emp.data_devolucao_prevista, '%d/%m/%Y').strftime('%Y-%m-%dT00:00')
            except:
                pass
        
        # ========== LÓGICA DE CORES E PRORROGAÇÃO ==========
        cor_fundo = ''
        cor_fonte = ''
        fonte_negrito = False
        eh_prorrogacao = False
        dias_display = dias
        
        if emp.status == 'EMPRESTADO' and emp.data_devolucao_prevista:
            try:
                # Verificar se é prorrogação (data_devolucao preenchida)
                eh_prorrogacao = emp.data_devolucao and emp.data_devolucao.strip() != ''
                
                # Converter data prevista
                data_prev_str = emp.data_devolucao_prevista.split(' ')[0] if ' ' in emp.data_devolucao_prevista else emp.data_devolucao_prevista
                data_prevista = datetime.strptime(data_prev_str, '%d/%m/%Y').date()
                
                # Calcular dias restantes
                dias_restantes = (data_prevista - hoje_date).days
                
                if dias_restantes < 0:
                    # VENCIDO - Vermelho com fonte vermelha negrito
                    cor_fundo = '#ffcccc'
                    cor_fonte = '#dc3545'
                    fonte_negrito = True
                    dias_display = f"VENCIDO ({abs(dias_restantes)} dias)"
                elif dias_restantes == 0:
                    # VENCE HOJE - Branco com fonte vermelha negrito
                    cor_fundo = 'white'
                    cor_fonte = '#dc3545'
                    fonte_negrito = True
                    dias_display = 'VENCE HOJE'
                elif dias_restantes == 1:
                    # FALTA 1 DIA - Vermelho clarinho
                    cor_fundo = '#ffe6e6'
                    cor_fonte = 'black'
                    fonte_negrito = False
                    dias_display = '1 dia (vence amanhã)'
                elif eh_prorrogacao:
                    # PRORROGAÇÃO - Verde clarinho
                    cor_fundo = '#e6ffe6'
                    cor_fonte = 'black'
                    fonte_negrito = False
                    dias_display = f"{dias_restantes} dias restantes (prorrogado)"
                else:
                    # NORMAL - Sem cor especial
                    cor_fundo = ''
                    cor_fonte = 'black'
                    fonte_negrito = False
                    dias_display = f"{dias_restantes} dias restantes"
            except Exception as e:
                print(f"Erro ao calcular cores para empréstimo {emp.id}: {str(e)}")
                import traceback
                traceback.print_exc()
        
        resultado.append({
            'id': emp.id,
            'codigo': emp.codigo,
            'data': emp.data,
            'data_iso': data_iso,
            'funcionario': emp.funcionario,
            'cargo': emp.cargo,
            'cod_produto': emp.cod_produto,
            'descricao_item': emp.descricao_item,
            'quantidade': grupo['quantidade_total'],  # QUANTIDADE AGRUPADA
            'observacoes': emp.observacoes,
            'status': emp.status,
            'data_devolucao_prevista': emp.data_devolucao_prevista,
            'data_devolucao_prevista_iso': data_devolucao_prevista_iso,
            'data_devolucao': emp.data_devolucao,
            'data_devolucao_iso': data_devolucao_iso,
            'dias': dias_display,
            'cor_fundo': cor_fundo,
            'cor_fonte': cor_fonte,
            'fonte_negrito': fonte_negrito,
            'eh_prorrogacao': eh_prorrogacao
        })
    
    return jsonify(resultado)

@app.route('/api/emprestimos', methods=['POST'])
def api_emprestimos_criar():
    data = request.json
    
    cod_produto = data.get('cod_produto')
    
    # VALIDAÇÃO CRÍTICA: Verificar estoque disponível
    if cod_produto:
        produto = Produto.query.get(cod_produto)
        if produto:
            estoque_atual = produto.estoque_atual
            
            # Somar quantidades emprestadas
            emprestimos_ativos = Emprestimo.query.filter_by(
                cod_produto=cod_produto,
                status='EMPRESTADO'
            ).all()
            emprestados = sum(e.quantidade or 1 for e in emprestimos_ativos)
            
            # Calcular disponível
            disponivel = estoque_atual - emprestados
            
            if disponivel <= 0:
                return jsonify({
                    'success': False,
                    'message': f'❌ ESTOQUE INSUFICIENTE!\n\nEstoque Total: {estoque_atual}\nJá Emprestado: {emprestados}\nDisponível: {disponivel}\n\nNão é possível emprestar este item!'
                })
    
    # Converter data de datetime-local (2026-01-21T14:30) para formato brasileiro com hora
    data_emp = data.get('data', '')
    if data_emp:
        try:
            # Formato: 2026-01-21T14:30 -> 21/01/2026 14:30
            dt = datetime.strptime(data_emp, '%Y-%m-%dT%H:%M')
            data_emp = dt.strftime('%d/%m/%Y %H:%M')
        except:
            # Fallback para data sem hora
            try:
                dt = datetime.strptime(data_emp, '%Y-%m-%d')
                data_emp = dt.strftime('%d/%m/%Y %H:%M')
            except:
                data_emp = datetime.now().strftime('%d/%m/%Y %H:%M')
    else:
        data_emp = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # Converter data_devolucao_prevista
    data_dev_prevista = data.get('data_devolucao_prevista', '')
    if data_dev_prevista:
        try:
            dt = datetime.strptime(data_dev_prevista, '%Y-%m-%dT%H:%M')
            data_dev_prevista = dt.strftime('%d/%m/%Y %H:%M')
        except:
            try:
                dt = datetime.strptime(data_dev_prevista, '%Y-%m-%d')
                data_dev_prevista = dt.strftime('%d/%m/%Y %H:%M')
            except:
                data_dev_prevista = None
    else:
        data_dev_prevista = None
    
    # Gerar código
    total = Emprestimo.query.count()
    codigo = str(total + 1)
    
    emprestimo = Emprestimo(
        codigo=codigo,
        data=data_emp,
        funcionario=data.get('funcionario', ''),
        cargo=data.get('cargo', ''),
        cod_produto=cod_produto,
        descricao_item=data.get('descricao_item', ''),
        quantidade=data.get('quantidade', 1),
        observacoes=data.get('observacoes', ''),
        status='EMPRESTADO',
        data_devolucao_prevista=data_dev_prevista
    )
    
    db.session.add(emprestimo)
    db.session.commit()
    
    return jsonify({'success': True, 'id': emprestimo.id})

@app.route('/api/emprestimos/<int:id>', methods=['PUT'])
def api_emprestimos_atualizar(id):
    emprestimo = Emprestimo.query.get_or_404(id)
    data = request.json
    
    if 'status' in data:
        emprestimo.status = data['status']
    if 'data_devolucao' in data:
        # Usar a hora atual do servidor ao invés da hora enviada pelo cliente
        emprestimo.data_devolucao = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/emprestimos/<int:id>/prorrogar', methods=['PUT'])
def api_emprestimos_prorrogar(id):
    """Prorrogar um empréstimo alterando a data de devolução prevista"""
    emprestimo = Emprestimo.query.get_or_404(id)
    data = request.json
    
    if 'data_devolucao_prevista' not in data:
        return jsonify({'success': False, 'message': 'Data de devolução prevista não informada'})
    
    nova_data = data['data_devolucao_prevista']
    
    # Converter data de datetime-local (2026-01-21T14:30) para formato brasileiro
    try:
        dt = datetime.strptime(nova_data, '%Y-%m-%dT%H:%M')
        nova_data = dt.strftime('%d/%m/%Y %H:%M')
    except:
        try:
            dt = datetime.strptime(nova_data, '%Y-%m-%d')
            nova_data = dt.strftime('%d/%m/%Y %H:%M')
        except:
            return jsonify({'success': False, 'message': 'Formato de data inválido'})
    
    print(f"[PRORROGAR] Nova data convertida: {nova_data}")
    
    # Registrar a prorrogação preenchendo data_devolucao com a data anterior
    # Isso marca que houve uma prorrogação
    if not emprestimo.data_devolucao:
        emprestimo.data_devolucao = emprestimo.data_devolucao_prevista
        print(f"[PRORROGAR] Salvando data anterior em data_devolucao: {emprestimo.data_devolucao}")
    
    # Atualizar a nova data prevista
    emprestimo.data_devolucao_prevista = nova_data
    
    db.session.commit()
    print(f"[PRORROGAR] Empréstimo atualizado com sucesso")
    return jsonify({'success': True})

@app.route('/api/emprestimos/<int:id>/alterar', methods=['PUT'])
def api_emprestimos_alterar_dados(id):
    emprestimo = Emprestimo.query.get_or_404(id)
    data = request.json
    
    if 'funcionario' in data:
        emprestimo.funcionario = data['funcionario']
    if 'observacoes' in data:
        emprestimo.observacoes = data['observacoes']
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/emprestimos/<int:id>', methods=['DELETE'])
def api_emprestimos_deletar(id):
    emprestimo = Emprestimo.query.get_or_404(id)
    db.session.delete(emprestimo)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/exportar-emprestimos-excel')
def exportar_emprestimos_excel():
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    # Buscar configurações
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Aplicar filtros
    data_inicial = request.args.get('data_inicial')
    data_final = request.args.get('data_final')
    descricao = request.args.get('descricao')
    status = request.args.get('status')
    funcionario = request.args.get('funcionario')
    observacao = request.args.get('observacao', '').lower()
    
    # Query base
    query = Emprestimo.query
    
    # Aplicar filtros
    if data_inicial:
        query = query.filter(Emprestimo.data >= data_inicial)
    if data_final:
        query = query.filter(Emprestimo.data <= data_final)
    if descricao:
        query = query.filter(Emprestimo.descricao_item.ilike(f'%{descricao}%'))
    if status:
        query = query.filter(Emprestimo.status == status)
    if funcionario:
        query = query.filter(Emprestimo.funcionario.ilike(f'%{funcionario}%'))
    
    emprestimos = query.order_by(Emprestimo.data.desc()).all()
    
    # Filtrar por observação
    if observacao:
        emprestimos = [e for e in emprestimos if observacao in (e.observacoes or '').lower()]
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empréstimos"
    
    current_row = 1
    
    # LOGO E TÍTULO
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass
    
    # Nome da empresa
    ws.merge_cells(f'B1:H3')
    title_cell = ws['B1']
    title_cell.value = nome_empresa
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 5
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Cabeçalhos
    headers = ['Data Lançamento', 'Funcionário', 'Descrição do Item', 'Status', 'Data Devolução', 'Dias', 'Observação']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Dados
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_font = Font(size=10)
    
    current_row += 1
    
    hoje = datetime.now().date()
    
    for emp in emprestimos:
        data_formatada = emp.data if emp.data else ''
        status_emp = emp.status
        data_dev = emp.data_devolucao if emp.data_devolucao else '---'
        dias = '---'
        
        # Calcular dias
        if emp.status == 'EMPRESTADO' and emp.data_devolucao_prevista:
            try:
                data_prevista = datetime.strptime(emp.data_devolucao_prevista.split(' ')[0], '%d/%m/%Y').date()
                diff = (data_prevista - hoje).days
                
                if diff < 0:
                    dias = f'VENCIDO ({abs(diff)} dias)'
                elif diff == 0:
                    dias = 'VENCE HOJE'
                elif diff == 1:
                    dias = '1 dia (vence amanhã)'
                else:
                    dias = f'{diff} dias restantes'
                
                data_dev = emp.data_devolucao_prevista
            except:
                pass
        
        row_data = [
            data_formatada,
            emp.funcionario,
            emp.descricao_item,
            status_emp,
            data_dev,
            dias,
            emp.observacoes or ''
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border_style
        
        current_row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 40
    
    # RODAPÉ
    if rodape:
        current_row += 2
        ws.merge_cells(f'A{current_row}:G{current_row}')
        rodape_cell = ws[f'A{current_row}']
        rodape_cell.value = rodape
        rodape_cell.font = Font(italic=True, size=10, color="666666")
        rodape_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Informação de geração
    current_row += 1
    ws.merge_cells(f'A{current_row}:G{current_row}')
    info_cell = ws[f'A{current_row}']
    info_cell.value = f'{len(emprestimos)} registro(s) - Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    info_cell.font = Font(italic=True, size=9, color="999999")
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'emprestimos_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/exportar-emprestimos-pdf')
def exportar_emprestimos_pdf():
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import os
    
    # Buscar configurações
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Aplicar filtros
    data_inicial = request.args.get('data_inicial')
    data_final = request.args.get('data_final')
    descricao = request.args.get('descricao')
    status = request.args.get('status')
    funcionario = request.args.get('funcionario')
    observacao = request.args.get('observacao', '').lower()
    
    # Query base
    query = Emprestimo.query
    
    # Aplicar filtros
    if data_inicial:
        query = query.filter(Emprestimo.data >= data_inicial)
    if data_final:
        query = query.filter(Emprestimo.data <= data_final)
    if descricao:
        query = query.filter(Emprestimo.descricao_item.ilike(f'%{descricao}%'))
    if status:
        query = query.filter(Emprestimo.status == status)
    if funcionario:
        query = query.filter(Emprestimo.funcionario.ilike(f'%{funcionario}%'))
    
    emprestimos = query.order_by(Emprestimo.data.desc()).all()
    
    # Filtrar por observação
    if observacao:
        emprestimos = [e for e in emprestimos if observacao in (e.observacoes or '').lower()]
    
    # Criar PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Logo e título
    if logo_path:
        try:
            logo = RLImage(logo_path, width=40*mm, height=20*mm)
            elements.append(logo)
            elements.append(Spacer(1, 5*mm))
        except:
            pass
    
    # Título
    elements.append(Paragraph(nome_empresa, title_style))
    elements.append(Paragraph("Relatório de Empréstimos", styles['Heading2']))
    elements.append(Spacer(1, 5*mm))
    
    # Tabela
    hoje = datetime.now().date()
    
    data = [['Data', 'Funcionário', 'Descrição', 'Status', 'Devolução', 'Dias', 'Observação']]
    
    for emp in emprestimos:
        data_formatada = emp.data if emp.data else ''
        status_emp = emp.status
        data_dev = emp.data_devolucao if emp.data_devolucao else '---'
        dias = '---'
        
        # Calcular dias
        if emp.status == 'EMPRESTADO' and emp.data_devolucao_prevista:
            try:
                data_prevista = datetime.strptime(emp.data_devolucao_prevista.split(' ')[0], '%d/%m/%Y').date()
                diff = (data_prevista - hoje).days
                
                if diff < 0:
                    dias = f'VENCIDO ({abs(diff)}d)'
                elif diff == 0:
                    dias = 'VENCE HOJE'
                elif diff == 1:
                    dias = '1 dia'
                else:
                    dias = f'{diff} dias'
                
                data_dev = emp.data_devolucao_prevista
            except:
                pass
        
        data.append([
            data_formatada,
            emp.funcionario[:20],
            emp.descricao_item[:30],
            status_emp,
            data_dev,
            dias,
            (emp.observacoes or '')[:25]
        ])
    
    table = Table(data, colWidths=[35*mm, 45*mm, 60*mm, 25*mm, 35*mm, 30*mm, 45*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 10*mm))
    
    # Rodapé
    if rodape:
        rodape_style = ParagraphStyle('Rodape', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER, fontName='Helvetica-Oblique')
        elements.append(Paragraph(rodape, rodape_style))
    
    # Info de geração
    info_text = f'{len(emprestimos)} registro(s) - Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph(info_text, info_style))
    
    # Gerar PDF
    doc.build(elements)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'emprestimos_{timestamp}.pdf'
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/emprestimos/multiplo', methods=['POST'])
def api_emprestimos_multiplo():
    data = request.json
    
    # Converter data de ISO para formato brasileiro
    data_emp = data.get('data', '')
    if data_emp:
        try:
            data_emp = datetime.strptime(data_emp, '%Y-%m-%d').strftime('%d/%m/%Y')
        except:
            pass
    
    funcionario = data.get('funcionario', '')
    observacao = data.get('observacao', '')
    itens = data.get('itens', [])
    
    if not funcionario or not itens:
        return jsonify({'success': False, 'message': 'Dados incompletos'})
    
    try:
        # Buscar cargo do funcionário
        func = Funcionario.query.filter_by(nome=funcionario).first()
        cargo = func.cargo if func else ''
        
        total_cadastrados = 0
        
        for item in itens:
            # Para cada quantidade, criar um registro individual
            for i in range(item['quantidade']):
                # Gerar código
                total = Emprestimo.query.count()
                codigo = str(total + 1)
                
                emprestimo = Emprestimo(
                    codigo=codigo,
                    data=data_emp,
                    funcionario=funcionario,
                    cargo=cargo,
                    cod_produto=item['codigo'],
                    descricao_item=item['descricao'],
                    observacoes=observacao,
                    status='EMPRESTADO'
                )
                
                db.session.add(emprestimo)
                total_cadastrados += 1
        
        db.session.commit()
        
        return jsonify({'success': True, 'total': total_cadastrados})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})



# APIs para Unidades
@app.route('/api/unidades', methods=['GET'])
def api_unidades_listar():
    unidades = Unidade.query.order_by(Unidade.nome).all()
    return jsonify([{'id': u.id, 'nome': u.nome} for u in unidades])

@app.route('/api/unidades', methods=['POST'])
def api_unidades_criar():
    data = request.json
    nome = data['nome'].strip()
    
    # Buscar TODAS as unidades e comparar em Python (case-insensitive)
    todas = Unidade.query.all()
    for uni in todas:
        if uni.nome.lower() == nome.lower():
            return jsonify({'success': False, 'message': f'Esta unidade já existe como "{uni.nome}"!'})
    
    try:
        unidade = Unidade(nome=nome)
        db.session.add(unidade)
        db.session.commit()
        return jsonify({'success': True, 'id': unidade.id})
    except:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erro ao cadastrar unidade'})

@app.route('/api/unidades/<int:id>', methods=['DELETE'])
def api_unidades_deletar(id):
    unidade = Unidade.query.get_or_404(id)
    db.session.delete(unidade)
    db.session.commit()
    return jsonify({'success': True})

# APIs para Marcas
@app.route('/api/marcas', methods=['GET'])
def api_marcas_listar():
    marcas = Marca.query.order_by(Marca.nome).all()
    return jsonify([{'id': m.id, 'nome': m.nome} for m in marcas])

@app.route('/api/marcas', methods=['POST'])
def api_marcas_criar():
    data = request.json
    nome = data['nome'].strip()
    
    # Buscar TODAS as marcas e comparar em Python (case-insensitive)
    todas = Marca.query.all()
    for marca in todas:
        if marca.nome.lower() == nome.lower():
            return jsonify({'success': False, 'message': f'Esta marca já existe como "{marca.nome}"!'})
    
    try:
        marca = Marca(nome=nome)
        db.session.add(marca)
        db.session.commit()
        return jsonify({'success': True, 'id': marca.id})
    except:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erro ao cadastrar marca'})

@app.route('/api/marcas/<int:id>', methods=['DELETE'])
def api_marcas_deletar(id):
    marca = Marca.query.get_or_404(id)
    db.session.delete(marca)
    db.session.commit()
    return jsonify({'success': True})

# APIs para Categorias
@app.route('/api/categorias', methods=['GET'])
def api_categorias_listar():
    categorias = Categoria.query.order_by(Categoria.nome).all()
    return jsonify([{'id': c.id, 'nome': c.nome} for c in categorias])

@app.route('/api/categorias', methods=['POST'])
def api_categorias_criar():
    data = request.json
    nome = data['nome'].strip()
    
    # Buscar TODAS as categorias e comparar em Python (case-insensitive)
    todas = Categoria.query.all()
    for cat in todas:
        if cat.nome.lower() == nome.lower():
            return jsonify({'success': False, 'message': f'Esta categoria já existe como "{cat.nome}"!'})
    
    try:
        categoria = Categoria(nome=nome)
        db.session.add(categoria)
        db.session.commit()
        return jsonify({'success': True, 'id': categoria.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erro ao cadastrar categoria'})

@app.route('/api/categorias/<int:id>', methods=['DELETE'])
def api_categorias_deletar(id):
    categoria = Categoria.query.get_or_404(id)
    db.session.delete(categoria)
    db.session.commit()
    return jsonify({'success': True})

# API para Ranking de Fornecedores
@app.route('/api/ranking-fornecedores', methods=['GET'])
def api_ranking_fornecedores():
    """
    Retorna ranking de fornecedores por gastos
    Filtros: categoria, ano
    """
    try:
        categoria = request.args.get('categoria', '')
        ano = request.args.get('ano', '')
        
        # Query base: produtos com fornecedor
        query = db.session.query(
            Produto.fornecedor,
            db.func.sum(Produto.preco_compra * Produto.estoque_atual).label('total_gasto')
        ).filter(
            Produto.fornecedor.isnot(None),
            Produto.fornecedor != '',
            Produto.ativo == True
        )
        
        # Filtrar por categoria se especificado
        if categoria:
            query = query.filter(Produto.categoria == categoria)
        
        # Filtrar por ano se especificado (usando data_cadastro)
        if ano:
            from sqlalchemy import extract
            query = query.filter(extract('year', Produto.data_cadastro) == int(ano))
        
        # Agrupar por fornecedor e ordenar por total gasto (decrescente)
        ranking = query.group_by(Produto.fornecedor)\
                      .order_by(db.desc('total_gasto'))\
                      .limit(10)\
                      .all()
        
        # Debug: imprimir resultados
        print(f"DEBUG: Encontrados {len(ranking)} fornecedores")
        for r in ranking:
            print(f"  - {r.fornecedor}: R$ {r.total_gasto}")
        
        # Calcular total geral
        total_gasto = sum([r.total_gasto if r.total_gasto else 0 for r in ranking])
        print(f"DEBUG: Total gasto = R$ {total_gasto}")
        
        # Formatar resultado
        resultado = {
            'ranking': [
                {
                    'fornecedor': r.fornecedor,
                    'total': float(r.total_gasto) if r.total_gasto else 0
                }
                for r in ranking
            ],
            'total_gasto': float(total_gasto) if total_gasto else 0
        }
        
        return jsonify(resultado)
        
    except Exception as e:
        import traceback
        print(f"Erro no ranking: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'ranking': [],
            'total_gasto': 0,
            'error': str(e)
        }), 500

# APIs para Operações
@app.route('/api/operacoes', methods=['GET'])
def api_operacoes_listar():
    operacoes = Operacao.query.order_by(Operacao.nome).all()
    return jsonify([{'id': o.id, 'nome': o.nome} for o in operacoes])

@app.route('/api/operacoes', methods=['POST'])
def api_operacoes_criar():
    data = request.json
    nome = data['nome'].strip()
    
    # Buscar TODAS as operações e comparar em Python (case-insensitive)
    todas = Operacao.query.all()
    for op in todas:
        if op.nome.lower() == nome.lower():
            return jsonify({'success': False, 'message': f'Esta operação já existe como "{op.nome}"!'})
    
    try:
        operacao = Operacao(nome=nome)
        db.session.add(operacao)
        db.session.commit()
        return jsonify({'success': True, 'id': operacao.id})
    except:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erro ao cadastrar operação'})

@app.route('/api/operacoes/<int:id>', methods=['DELETE'])
def api_operacoes_deletar(id):
    operacao = Operacao.query.get_or_404(id)
    db.session.delete(operacao)
    db.session.commit()
    return jsonify({'success': True})

# APIs para Anos
@app.route('/api/anos', methods=['GET'])
def api_anos_listar():
    anos = Ano.query.order_by(Ano.ano.desc()).all()
    return jsonify([{'id': a.id, 'ano': a.ano} for a in anos])

@app.route('/api/anos', methods=['POST'])
def api_anos_criar():
    data = request.json
    ano = int(data['ano'])
    
    # Verificar se o ano já existe
    existe = Ano.query.filter_by(ano=ano).first()
    if existe:
        return jsonify({'success': False, 'message': f'O ano {ano} já está cadastrado!'})
    
    try:
        novo_ano = Ano(ano=ano)
        db.session.add(novo_ano)
        db.session.commit()
        return jsonify({'success': True, 'id': novo_ano.id})
    except:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erro ao cadastrar ano'})

@app.route('/api/anos/<int:id>', methods=['DELETE'])
def api_anos_deletar(id):
    ano = Ano.query.get_or_404(id)
    db.session.delete(ano)
    db.session.commit()
    return jsonify({'success': True})

# ==================== EXPORTAÇÃO EXCEL ====================

@app.route('/api/exportar-categoria-excel')
def exportar_categoria_excel():
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    # Buscar configurações (com tratamento de erro)
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        # Se a tabela não existir, usar valores padrão
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Parâmetros (aplicar filtros)
    categoria = request.args.get('categoria', '')
    status = request.args.get('status', '')
    pesquisa = request.args.get('pesquisa', '')
    
    # Buscar produtos
    query = Produto.query.filter_by(ativo=True)
    
    if categoria:
        query = query.filter_by(categoria=categoria)
    
    produtos = query.all()
    
    # Filtrar por status
    if status == 'baixo':
        produtos = [p for p in produtos if p.estoque_atual < p.estoque_minimo]
    elif status == 'excedido':
        produtos = [p for p in produtos if p.estoque_atual > p.estoque_maximo]
    elif status == 'normal':
        produtos = [p for p in produtos if p.estoque_minimo <= p.estoque_atual <= p.estoque_maximo]
    
    # Filtrar por pesquisa
    if pesquisa:
        produtos = [p for p in produtos if 
                   pesquisa.lower() in (p.descricao or '').lower() or
                   pesquisa.lower() in (p.marca or '').lower() or
                   pesquisa.lower() in (p.observacao or '').lower()]
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categoria Para Cotação"
    
    current_row = 1
    
    # LOGO E TÍTULO
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass
    
    # Nome da empresa (centralizado)
    ws.merge_cells(f'B1:H3')
    title_cell = ws['B1']
    title_cell.value = nome_empresa
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 5
    
    # Estilo do cabeçalho da tabela
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Cabeçalhos
    headers = ['Data Atual', 'Descrição do Item', 'Marca', 'Qtd Entrada', 
               'Saldo do Estoque', 'Estoque Mínimo', 'Estoque Máximo', 'Empréstimo']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Dados com bordas
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_font = Font(size=10)
    
    current_row += 1
    data_atual = datetime.now().strftime('%d/%m/%Y')
    
    for produto in produtos:
        # Buscar movimentações para calcular total de entradas
        movimentacoes = Movimentacao.query.filter_by(produto_codigo=produto.codigo).all()
        total_entradas = sum(m.quantidade for m in movimentacoes if m.tipo == 'ENTRADA')
        
        # Buscar empréstimos ativos (usando cod_produto e status EMPRESTADO)
        # SOMA as quantidades de todos os empréstimos ativos
        emprestimos_ativos = Emprestimo.query.filter_by(cod_produto=produto.codigo, status='EMPRESTADO').all()
        total_emprestimos = sum(e.quantidade for e in emprestimos_ativos)
        
        row_data = [
            data_atual,
            produto.descricao or '',
            produto.marca or '',
            total_entradas,
            produto.estoque_atual or 0,
            produto.estoque_minimo or 0,
            produto.estoque_maximo or 0,
            total_emprestimos
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border_style
        
        current_row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    # RODAPÉ
    if rodape:
        current_row += 2
        ws.merge_cells(f'A{current_row}:H{current_row}')
        rodape_cell = ws[f'A{current_row}']
        rodape_cell.value = rodape
        rodape_cell.font = Font(italic=True, size=10, color="666666")
        rodape_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar informação de geração
    current_row += 1
    ws.merge_cells(f'A{current_row}:H{current_row}')
    info_cell = ws[f'A{current_row}']
    info_cell.value = f'{len(produtos)} produto(s) - Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    info_cell.font = Font(italic=True, size=9, color="999999")
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'categoria_cotacao_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/exportar-categoria-pdf')
def exportar_categoria_pdf():
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import os
    
    # Buscar configurações (com tratamento de erro)
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        # Se a tabela não existir, usar valores padrão
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Parâmetros (aplicar filtros)
    categoria = request.args.get('categoria', '')
    status = request.args.get('status', '')
    pesquisa = request.args.get('pesquisa', '')
    
    # Buscar produtos
    query = Produto.query.filter_by(ativo=True)
    
    if categoria:
        query = query.filter_by(categoria=categoria)
    
    produtos = query.all()
    
    # Filtrar por status
    if status == 'baixo':
        produtos = [p for p in produtos if p.estoque_atual < p.estoque_minimo]
    elif status == 'excedido':
        produtos = [p for p in produtos if p.estoque_atual > p.estoque_maximo]
    elif status == 'normal':
        produtos = [p for p in produtos if p.estoque_minimo <= p.estoque_atual <= p.estoque_maximo]
    
    # Filtrar por pesquisa
    if pesquisa:
        produtos = [p for p in produtos if 
                   pesquisa.lower() in (p.descricao or '').lower() or
                   pesquisa.lower() in (p.marca or '').lower() or
                   pesquisa.lower() in (p.observacao or '').lower()]
    
    # Criar PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2c3e50'), spaceAfter=12, alignment=TA_CENTER, fontName='Helvetica-Bold')
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'), alignment=TA_CENTER, fontName='Helvetica-Oblique')
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#999999'), alignment=TA_CENTER, fontName='Helvetica-Oblique')
    
    # Elementos do PDF
    elements = []
    
    # LOGO E TÍTULO
    if logo_path:
        try:
            logo = RLImage(logo_path, width=3*cm, height=1.5*cm)
            # Criar tabela para logo e título lado a lado
            header_data = [[logo, Paragraph(nome_empresa, title_style)]]
            header_table = Table(header_data, colWidths=[4*cm, 20*cm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(header_table)
        except:
            elements.append(Paragraph(nome_empresa, title_style))
    else:
        elements.append(Paragraph(nome_empresa, title_style))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Dados da tabela
    data_atual = datetime.now().strftime('%d/%m/%Y')
    data = [['Data Atual', 'Descrição do Item', 'Marca', 'Qtd Entrada', 
             'Saldo do Estoque', 'Estoque Mínimo', 'Estoque Máximo', 'Empréstimo']]
    
    for produto in produtos:
        # Buscar movimentações para calcular total de entradas
        movimentacoes = Movimentacao.query.filter_by(produto_codigo=produto.codigo).all()
        total_entradas = sum(m.quantidade for m in movimentacoes if m.tipo == 'ENTRADA')
        
        # Buscar empréstimos ativos (usando cod_produto e status EMPRESTADO)
        # SOMA as quantidades de todos os empréstimos ativos
        emprestimos_ativos = Emprestimo.query.filter_by(cod_produto=produto.codigo, status='EMPRESTADO').all()
        total_emprestimos = sum(e.quantidade for e in emprestimos_ativos)
        
        data.append([
            data_atual,
            produto.descricao or '',
            produto.marca or '',
            str(total_entradas),
            str(produto.estoque_atual or 0),
            str(produto.estoque_minimo or 0),
            str(produto.estoque_maximo or 0),
            str(total_emprestimos)
        ])
    
    # Criar tabela
    table = Table(data, colWidths=[2.5*cm, 7*cm, 3.5*cm, 2.5*cm, 3*cm, 2.8*cm, 2.8*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))
    
    elements.append(table)
    
    # RODAPÉ
    if rodape:
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(rodape, footer_style))
    
    # Informação de geração
    elements.append(Spacer(1, 0.3*cm))
    info_text = f'{len(produtos)} produto(s) - Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    elements.append(Paragraph(info_text, info_style))
    
    # Gerar PDF
    doc.build(elements)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'categoria_cotacao_{timestamp}.pdf'
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ==================== APPLICATION STARTUP (MOVIDO PARA O FINAL) ====================

@app.route('/api/exportar-movimentacoes-excel')
def exportar_movimentacoes_excel():
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    # Buscar configurações (com tratamento de erro)
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        # Se a tabela não existir, usar valores padrão
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Aplicar os mesmos filtros da listagem
    data_inicial = request.args.get('data_inicial')
    data_final = request.args.get('data_final')
    tipo = request.args.get('tipo')
    mes = request.args.get('mes')
    nota_fiscal = request.args.get('nota_fiscal')
    
    # Query base
    query = Movimentacao.query
    
    # Aplicar filtros
    if data_inicial:
        query = query.filter(Movimentacao.data_movimentacao >= data_inicial)
    if data_final:
        query = query.filter(Movimentacao.data_movimentacao <= data_final)
    if tipo and tipo != 'TODOS':
        query = query.filter(Movimentacao.tipo == tipo)
    if mes and mes != 'TODOS':
        from sqlalchemy import extract
        query = query.filter(extract('month', Movimentacao.data_movimentacao) == int(mes))
    if nota_fiscal:
        query = query.filter(Movimentacao.nota_fiscal.like(f'%{nota_fiscal}%'))
    
    # Buscar movimentações filtradas
    movimentacoes = query.order_by(Movimentacao.data_movimentacao.desc()).all()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações"
    
    current_row = 1
    
    # LOGO E TÍTULO
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass
    
    # Nome da empresa (centralizado)
    ws.merge_cells(f'B1:I3')
    title_cell = ws['B1']
    title_cell.value = nome_empresa
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 5
    
    # Estilo do cabeçalho da tabela
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Cabeçalhos - TODAS AS COLUNAS
    headers = ['Data', 'Cód.Produto', 'Descrição', 'E/S', 'Quantidade', 'Solicitante', 'NF', 'Operação', 'Valor Total', 'Usuário']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Dados com bordas
    data_alignment = Alignment(horizontal="left", vertical="center")
    valor_alignment = Alignment(horizontal="right", vertical="center")
    data_font = Font(size=10)
    
    current_row += 1
    start_data_row = current_row
    
    for m in movimentacoes:
        produto = Produto.query.get(m.produto_codigo)
        data = m.data_movimentacao.strftime('%d/%m/%Y') if m.data_movimentacao else ''
        tipo = 'E' if m.tipo == 'ENTRADA' else ('S' if m.tipo == 'SAIDA' else 'D')
        
        # Buscar usuário real
        usuario_nome = 'Sistema'
        if m.usuario_id:
            usuario = Usuario.query.get(m.usuario_id)
            if usuario:
                usuario_nome = usuario.nome
        
        # Determinar solicitante
        solicitante = ''
        if m.tipo == 'ENTRADA':
            if produto and produto.fornecedor:
                solicitante = produto.fornecedor
            elif hasattr(m, 'solicitante_nome') and m.solicitante_nome:
                solicitante = m.solicitante_nome
        else:
            if hasattr(m, 'solicitante_nome') and m.solicitante_nome:
                solicitante = m.solicitante_nome
        
        # Calcular valor total
        valor_total = m.valor_total if m.valor_total else 0
        
        row_data = [
            data,
            m.produto_codigo,
            produto.descricao if produto else '',
            tipo,
            m.quantidade,
            solicitante,
            m.nota_fiscal or '',
            m.operacao or '',
            valor_total,
            usuario_nome
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            # Alinhar valor à direita
            if col == 9:  # Coluna Valor Total
                cell.alignment = valor_alignment
                cell.number_format = 'R$ #,##0.00'
            else:
                cell.alignment = data_alignment
            cell.border = border_style
        
        current_row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    # RODAPÉ
    if rodape:
        current_row += 2
        ws.merge_cells(f'A{current_row}:J{current_row}')
        rodape_cell = ws[f'A{current_row}']
        rodape_cell.value = rodape
        rodape_cell.font = Font(italic=True, size=10, color="666666")
        rodape_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar informação de geração
    current_row += 1
    ws.merge_cells(f'A{current_row}:J{current_row}')
    info_cell = ws[f'A{current_row}']
    info_cell.value = f'{len(movimentacoes)} registro(s) - Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    info_cell.font = Font(italic=True, size=9, color="999999")
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'movimentacoes_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/exportar-movimentacoes-pdf')
def exportar_movimentacoes_pdf():
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import os
    
    # Buscar configurações (com tratamento de erro)
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        # Se a tabela não existir, usar valores padrão
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Aplicar os mesmos filtros da listagem
    data_inicial = request.args.get('data_inicial')
    data_final = request.args.get('data_final')
    tipo = request.args.get('tipo')
    mes = request.args.get('mes')
    nota_fiscal = request.args.get('nota_fiscal')
    
    # Query base
    query = Movimentacao.query
    
    # Aplicar filtros
    if data_inicial:
        query = query.filter(Movimentacao.data_movimentacao >= data_inicial)
    if data_final:
        query = query.filter(Movimentacao.data_movimentacao <= data_final)
    if tipo and tipo != 'TODOS':
        query = query.filter(Movimentacao.tipo == tipo)
    if mes and mes != 'TODOS':
        from sqlalchemy import extract
        query = query.filter(extract('month', Movimentacao.data_movimentacao) == int(mes))
    if nota_fiscal:
        query = query.filter(Movimentacao.nota_fiscal.like(f'%{nota_fiscal}%'))
    
    # Buscar movimentações filtradas
    movimentacoes = query.order_by(Movimentacao.data_movimentacao.desc()).all()
    
    # Criar PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#2c3e50'), spaceAfter=12, alignment=TA_CENTER, fontName='Helvetica-Bold')
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'), alignment=TA_CENTER, fontName='Helvetica-Oblique')
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#999999'), alignment=TA_CENTER, fontName='Helvetica-Oblique')
    
    # Elementos do PDF
    elements = []
    
    # LOGO E TÍTULO
    if logo_path:
        try:
            logo = RLImage(logo_path, width=3*cm, height=1.5*cm)
            # Criar tabela para logo e título lado a lado
            header_data = [[logo, Paragraph(nome_empresa, title_style)]]
            header_table = Table(header_data, colWidths=[4*cm, 20*cm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(header_table)
        except:
            elements.append(Paragraph(nome_empresa, title_style))
    else:
        elements.append(Paragraph(nome_empresa, title_style))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Dados da tabela
    data = [['Data', 'Cód', 'Descrição', 'E/S', 'Qtd', 'Solicitante', 'NF', 'Operação', 'Usuário']]
    
    for m in movimentacoes:
        produto = Produto.query.get(m.produto_codigo)
        data_mov = m.data_movimentacao.strftime('%d/%m/%Y') if m.data_movimentacao else ''
        tipo = 'E' if m.tipo == 'ENTRADA' else ('S' if m.tipo == 'SAIDA' else 'D')
        
        # Buscar usuário real
        usuario_nome = 'Sistema'
        if m.usuario_id:
            usuario = Usuario.query.get(m.usuario_id)
            if usuario:
                usuario_nome = usuario.nome
        
        # Determinar solicitante
        solicitante = (m.observacao or m.operacao or '')[:15]
        
        data.append([
            data_mov,
            str(m.produto_codigo),
            (produto.descricao if produto else '')[:30],
            tipo,
            str(m.quantidade),
            solicitante,
            m.nota_fiscal or '',
            (m.operacao or '')[:15],
            usuario_nome[:15]
        ])
    
    # Criar tabela
    table = Table(data, colWidths=[2.5*cm, 1.5*cm, 6*cm, 1.5*cm, 1.5*cm, 3*cm, 2*cm, 3*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))
    
    # RODAPÉ
    if rodape:
        elements.append(Paragraph(rodape, footer_style))
        elements.append(Spacer(1, 0.3*cm))
    
    # Informação de geração
    info_text = f'{len(movimentacoes)} registro(s) - Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    elements.append(Paragraph(info_text, info_style))
    
    # Gerar PDF
    doc.build(elements)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'movimentacoes_{timestamp}.pdf'
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# ==================== SISTEMA - BACKUP E CONFIGURAÇÕES ====================

# Model para Configurações
class ConfiguracaoSistema(db.Model):
    __tablename__ = 'configuracoes_sistema'
    
    id = db.Column(db.Integer, primary_key=True)
    nome_empresa = db.Column(db.String(200), default='Minha Empresa')
    logo_path = db.Column(db.String(200))
    rodape = db.Column(db.String(500))
    data_atualizacao = db.Column(db.DateTime, default=datetime.now)

@app.route('/sistema')
def sistema():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    # Bloquear acesso para usuários intermediários
    if session.get('usuario_tipo') == 'intermediario':
        return redirect(url_for('index'))
    
    # Desabilitar cache
    response = make_response(render_template('sistema_novo.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/configuracoes/relatorios')
def configuracoes_relatorios():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    return render_template('configuracoes_relatorios.html')

# API para obter configurações
@app.route('/api/configuracoes/relatorios', methods=['GET'])
def get_configuracoes_relatorios():
    try:
        config = ConfiguracaoSistema.query.first()
        if not config:
            return jsonify({
                'success': True,
                'nome_empresa': '',
                'logo_path': '',
                'logo_url': '',
                'rodape': ''
            })
        
        logo_url = ''
        if config.logo_path:
            logo_url = url_for('static', filename=f'logos/{config.logo_path}')
        
        return jsonify({
            'success': True,
            'nome_empresa': config.nome_empresa or '',
            'logo_path': config.logo_path or '',
            'logo_url': logo_url,
            'rodape': config.rodape or ''
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# API para salvar configurações
@app.route('/api/configuracoes/relatorios', methods=['POST'])
def save_configuracoes_relatorios():
    try:
        nome_empresa = request.form.get('nome_empresa', '').strip()
        rodape = request.form.get('rodape', '').strip()
        
        config = ConfiguracaoSistema.query.first()
        
        if not config:
            config = ConfiguracaoSistema()
            db.session.add(config)
        
        config.nome_empresa = nome_empresa
        config.rodape = rodape
        config.data_atualizacao = datetime.now()
        
        # Upload do logo
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Criar pasta se não existir
                logos_dir = os.path.join(app.static_folder, 'logos')
                os.makedirs(logos_dir, exist_ok=True)
                
                # Gerar nome único
                ext = os.path.splitext(logo_file.filename)[1]
                filename = f'logo_empresa_{datetime.now().strftime("%Y%m%d_%H%M%S")}{ext}'
                filepath = os.path.join(logos_dir, filename)
                
                # Salvar arquivo
                logo_file.save(filepath)
                config.logo_path = filename
                print(f"Logo salvo: {filename}")
        
        print("Fazendo commit...")
        db.session.commit()
        print("✅ Commit realizado com sucesso!")
        return jsonify({'success': True, 'message': 'Configurações salvas com sucesso!'})
    except Exception as e:
        print(f"❌ ERRO: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# Backup Manual
@app.route('/api/backup/download', methods=['GET'])
def backup_download():
    import shutil
    import zipfile
    from io import BytesIO
    
    try:
        # Criar arquivo ZIP em memória
        memory_file = BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Adicionar banco de dados
            if os.path.exists('instance/estoque_web.db'):
                zipf.write('instance/estoque_web.db', 'estoque_web.db')
            
            # Adicionar pasta de fotos
            if os.path.exists('estoque_web/static/uploads/produtos'):
                for root, dirs, files in os.walk('estoque_web/static/uploads/produtos'):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.join('uploads', file)
                        zipf.write(file_path, arcname)
        
        memory_file.seek(0)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'backup_completo_{timestamp}.zip'
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/backup/restaurar', methods=['POST'])
def backup_restaurar():
    import shutil
    import zipfile
    
    try:
        if 'backup' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['backup']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.endswith('.zip'):
            return jsonify({'success': False, 'message': 'Arquivo deve ser .zip'}), 400
        
        # Salvar arquivo temporário
        temp_path = 'temp_backup.zip'
        file.save(temp_path)
        
        # Extrair ZIP
        with zipfile.ZipFile(temp_path, 'r') as zipf:
            # Restaurar banco de dados
            if 'estoque_web.db' in zipf.namelist():
                zipf.extract('estoque_web.db', 'temp_restore')
                shutil.copy2('temp_restore/estoque_web.db', 'instance/estoque_web.db')
            
            # Restaurar fotos
            upload_dir = 'estoque_web/static/uploads/produtos'
            os.makedirs(upload_dir, exist_ok=True)
            
            for file_info in zipf.namelist():
                if file_info.startswith('uploads/'):
                    zipf.extract(file_info, 'temp_restore')
                    filename = os.path.basename(file_info)
                    if filename:
                        shutil.copy2(
                            os.path.join('temp_restore', file_info),
                            os.path.join(upload_dir, filename)
                        )
        
        # Limpar arquivos temporários
        os.remove(temp_path)
        if os.path.exists('temp_restore'):
            shutil.rmtree('temp_restore')
        
        return jsonify({'success': True, 'message': 'Backup restaurado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
def backup_manual():
    try:
        import shutil
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"estoque_web_manual_{timestamp}.db"
        backup_path = os.path.join('Backup', 'ManualBackup', backup_filename)
        
        os.makedirs(os.path.dirname(backup_path), exist_ok=True)
        shutil.copy2('instance/estoque_web.db', backup_path)
        
        return jsonify({'success': True, 'filename': backup_filename})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Listar Backups
@app.route('/api/backup/listar', methods=['GET'])
def listar_backups():
    try:
        backups = []
        backup_dir = os.path.join('Backup', 'ManualBackup')
        if os.path.exists(backup_dir):
            for f in os.listdir(backup_dir):
                if f.endswith('.db'):
                    filepath = os.path.join(backup_dir, f)
                    backups.append({
                        'nome': f,
                        'tamanho': os.path.getsize(filepath),
                        'data': datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%d/%m/%Y %H:%M')
                    })
        return jsonify(backups)
    except Exception as e:
        return jsonify([])

# Configurações
@app.route('/api/configuracoes', methods=['GET'])
def get_configuracoes():
    config = ConfiguracaoSistema.query.first()
    if not config:
        config = ConfiguracaoSistema(nome_empresa='Minha Empresa')
        db.session.add(config)
        db.session.commit()
    
    return jsonify({
        'nome_empresa': config.nome_empresa,
        'logo_path': config.logo_path
    })

@app.route('/api/configuracoes', methods=['POST'])
def salvar_configuracoes():
    data = request.json
    config = ConfiguracaoSistema.query.first()
    if not config:
        config = ConfiguracaoSistema()
        db.session.add(config)
    
    config.nome_empresa = data.get('nome_empresa')
    config.logo_path = data.get('logo_path')
    config.data_atualizacao = datetime.now()
    
    db.session.commit()
    return jsonify({'success': True})

# Limpeza de Dados
@app.route('/api/limpeza/emprestimos', methods=['POST'])
def limpar_emprestimos():
    try:
        Emprestimo.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Empréstimos limpos com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/movimentacoes', methods=['POST'])
def limpar_movimentacoes():
    try:
        Movimentacao.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Movimentações limpas com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/produtos', methods=['POST'])
def limpar_produtos():
    try:
        # Apenas zerar as quantidades em estoque dos produtos
        produtos = Produto.query.all()
        for produto in produtos:
            produto.estoque_minimo = 0
            produto.estoque_maximo = 0
        
        # Deletar os produtos
        Produto.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Produtos e estoque limpos com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/funcionarios', methods=['POST'])
def limpar_funcionarios():
    try:
        Funcionario.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Funcionários limpos com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/fornecedores', methods=['POST'])
def limpar_fornecedores():
    try:
        Fornecedor.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Fornecedores limpos com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/clientes', methods=['POST'])
def limpar_clientes():
    try:
        Cliente.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Clientes limpos com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/estoque', methods=['POST'])
def zerar_estoque():
    try:
        # Zerar apenas o estoque_atual de todos os produtos
        produtos = Produto.query.all()
        for p in produtos:
            p.estoque_atual = 0
        db.session.commit()
        return jsonify({'success': True, 'message': 'Estoque zerado com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/limpeza/tudo', methods=['POST'])
def zerar_tudo():
    try:
        # Limpar todas as tabelas (exceto usuários, configurações e itens padrão)
        Emprestimo.query.delete()
        Movimentacao.query.delete()
        Produto.query.delete()
        Funcionario.query.delete()
        Fornecedor.query.delete()
        Cliente.query.delete()
        
        # NÃO deletar as tabelas de "Diversos" - manter os itens padrão
        # Unidade.query.delete()
        # Marca.query.delete()
        # Categoria.query.delete()
        # Operacao.query.delete()
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Todos os dados foram limpos com sucesso! (Itens padrão mantidos)'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/sistema/status', methods=['GET'])
def status_sistema():
    try:
        # Contar registros
        total_produtos = Produto.query.count()
        total_funcionarios = Funcionario.query.count()
        total_fornecedores = Fornecedor.query.count()
        total_clientes = Cliente.query.count()
        total_emprestimos = Emprestimo.query.filter_by(status='EMPRESTADO').count()
        total_movimentacoes = Movimentacao.query.count()
        
        # Tamanho do banco
        db_path = 'instance/estoque_web.db'
        db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        db_size_mb = db_size / 1024 / 1024
        
        return jsonify({
            'success': True,
            'produtos': total_produtos,
            'funcionarios': total_funcionarios,
            'fornecedores': total_fornecedores,
            'clientes': total_clientes,
            'emprestimos_ativos': total_emprestimos,
            'movimentacoes': total_movimentacoes,
            'tamanho_banco': f'{db_size_mb:.2f} MB'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/sistema/validar-banco', methods=['GET'])
def validar_banco():
    try:
        # Verificar integridade básica
        problemas = []
        
        # Verificar produtos sem estoque negativo
        produtos_negativos = Produto.query.filter(Produto.estoque_atual < 0).count()
        if produtos_negativos > 0:
            problemas.append(f'{produtos_negativos} produto(s) com estoque negativo')
        
        # Verificar empréstimos sem produto
        emprestimos = Emprestimo.query.all()
        emprestimos_invalidos = 0
        for emp in emprestimos:
            if emp.cod_produto:
                produto = Produto.query.get(emp.cod_produto)
                if not produto:
                    emprestimos_invalidos += 1
        
        if emprestimos_invalidos > 0:
            problemas.append(f'{emprestimos_invalidos} empréstimo(s) com produto inexistente')
        
        # Verificar movimentações sem produto
        movimentacoes = Movimentacao.query.all()
        movimentacoes_invalidas = 0
        for mov in movimentacoes:
            produto = Produto.query.get(mov.produto_codigo)
            if not produto:
                movimentacoes_invalidas += 1
        
        if movimentacoes_invalidas > 0:
            problemas.append(f'{movimentacoes_invalidas} movimentação(ões) com produto inexistente')
        
        if len(problemas) == 0:
            return jsonify({
                'success': True,
                'valido': True,
                'message': 'Banco de dados validado com sucesso! Nenhum problema encontrado.'
            })
        else:
            return jsonify({
                'success': True,
                'valido': False,
                'problemas': problemas
            })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ==================== EXPORTAR E IMPORTAR PRODUTOS ====================

# FUNÇÃO HELPER PADRÃO PARA TODAS AS EXPORTAÇÕES
def criar_excel_padrao(titulo_planilha, headers, dados, nome_arquivo):
    """
    Função padrão para criar arquivos Excel com formatação consistente.
    TODAS as exportações do sistema DEVEM usar esta função.
    
    Suporta formatos: .xlsx, .xlsm, .xlsb, .xls, .xltx, .xltm
    
    Args:
        titulo_planilha: Nome da aba do Excel
        headers: Lista com os nomes das colunas
        dados: Lista de listas com os dados (cada lista interna é uma linha)
        nome_arquivo: Nome do arquivo para download
    
    Returns:
        BytesIO com o arquivo Excel gerado
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    # Buscar configurações do sistema
    try:
        config = ConfiguracaoSistema.query.first()
        nome_empresa = config.nome_empresa if config and config.nome_empresa else 'Minha Empresa'
        rodape = config.rodape if config and config.rodape else ''
        logo_path = None
        if config and config.logo_path:
            logo_full_path = os.path.join(app.static_folder, 'logos', config.logo_path)
            if os.path.exists(logo_full_path):
                logo_path = logo_full_path
    except:
        nome_empresa = 'Minha Empresa'
        rodape = ''
        logo_path = None
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_planilha
    
    current_row = 1
    
    # LOGO (se existir)
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass
    
    # NOME DA EMPRESA (centralizado nas 3 primeiras linhas)
    num_colunas = len(headers)
    ultima_coluna = openpyxl.utils.get_column_letter(num_colunas)
    ws.merge_cells(f'B1:{ultima_coluna}3')
    title_cell = ws['B1']
    title_cell.value = nome_empresa
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 5
    
    # ESTILOS PADRÃO
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # CABEÇALHOS
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
    
    current_row += 1
    
    # DADOS
    for linha in dados:
        for col, valor in enumerate(linha, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = valor
            cell.border = border_style
        current_row += 1
    
    # RODAPÉ (se existir)
    if rodape:
        current_row += 2
        ws.merge_cells(f'A{current_row}:{ultima_coluna}{current_row}')
        rodape_cell = ws[f'A{current_row}']
        rodape_cell.value = rodape
        rodape_cell.font = Font(size=9, italic=True)
        rodape_cell.alignment = Alignment(horizontal='center')
    
    # AJUSTAR LARGURA DAS COLUNAS AUTOMATICAMENTE
    for col in range(1, num_colunas + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        max_length = 0
        for row in range(5, current_row):
            cell_value = ws[f'{col_letter}{row}'].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        # Definir largura (mínimo 10, máximo 50)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)
    
    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

@app.route('/api/exportar-produtos-excel')
def exportar_produtos_excel():
    from flask import send_file
    
    busca = request.args.get('busca', '')
    ativo = request.args.get('ativo', 'ativos')
    
    query = Produto.query
    
    if busca:
        query = query.filter(Produto.descricao.like(f'%{busca}%'))
    
    if ativo == 'ativos':
        query = query.filter(Produto.ativo == True)
    elif ativo == 'inativos':
        query = query.filter(Produto.ativo == False)
    
    produtos = query.order_by(Produto.descricao).all()
    
    # Preparar dados
    headers = ['Código', 'Descrição', 'Cód.Barras', 'Unidade', 'Marca', 'Categoria', 'Fornecedor', 'R$ Compra', 'R$ Venda', 'Est.Mín', 'Est.Máx', 'Status']
    dados = []
    
    for p in produtos:
        dados.append([
            p.codigo,
            p.descricao,
            p.codigo_barras or '',
            p.unidade or '',
            p.marca or '',
            p.categoria or '',
            p.fornecedor or '',
            p.preco_compra or 0,
            p.preco_venda or 0,
            p.estoque_minimo or 0,
            p.estoque_maximo or 0,
            'Ativo' if p.ativo else 'Inativo'
        ])
    
    # Usar função padrão
    output = criar_excel_padrao('Produtos', headers, dados, 'produtos.xlsx')
    
    return send_file(output, download_name='produtos.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/importar-produtos-excel', methods=['POST'])
def importar_produtos_excel():
    """Importar produtos em massa a partir de planilha Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use apenas .xlsx ou .xls'})
        
        # Ler o arquivo Excel
        df = pd.read_excel(arquivo, header=None)
        
        # Procurar a linha que contém "Descrição" ou "Código" (cabeçalhos reais)
        linha_cabecalho = 0
        for i in range(min(10, len(df))):
            linha_str = ' '.join([str(x).upper() for x in df.iloc[i].values if pd.notna(x)])
            if 'DESCRIÇÃO' in linha_str or 'DESCRICAO' in linha_str or 'CÓDIGO' in linha_str or 'CODIGO' in linha_str:
                linha_cabecalho = i
                break
        
        # Reler o arquivo com o cabeçalho correto
        df = pd.read_excel(arquivo, header=linha_cabecalho)
        
        # Remover colunas completamente vazias (Unnamed)
        df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed', na=False)]
        
        # Limpar nomes das colunas
        colunas_validas = []
        for col in df.columns:
            col_str = str(col).strip()
            if col_str and col_str.lower() not in ['nan', 'none', '']:
                colunas_validas.append(col)
        
        df = df[colunas_validas]
        df.columns = [str(col).strip() for col in df.columns]
        
        # Mapear colunas
        colunas_esperadas = {
            'codigo': ['Código', 'Codigo', 'CÓDIGO', 'CODIGO', 'Cód', 'COD', 'ID'],
            'descricao': ['Descrição', 'Descricao', 'DESCRIÇÃO', 'DESCRICAO', 'Produto', 'PRODUTO'],
            'codigo_barras': ['Cód.Barras', 'CÓD.BARRAS', 'Código de Barras', 'Cod.Barras'],
            'unidade': ['Unidade', 'UNIDADE', 'Un', 'UN'],
            'marca': ['Marca', 'MARCA'],
            'categoria': ['Categoria', 'CATEGORIA'],
            'fornecedor': ['Fornecedor', 'FORNECEDOR'],
            'preco_compra': ['R$ Compra', 'R$ COMPRA', 'Preço Compra', 'Compra'],
            'preco_venda': ['R$ Venda', 'R$ VENDA', 'Preço Venda', 'Venda'],
            'estoque_minimo': ['Est.Mín', 'EST.MÍN', 'Estoque Mínimo', 'Est.Min'],
            'estoque_maximo': ['Est.Máx', 'EST.MÁX', 'Estoque Máximo', 'Est.Max'],
            'status': ['Status', 'STATUS', 'Ativo', 'ATIVO'],
            'observacao': ['Observação', 'Observacao', 'Obs', 'OBS']
        }
        
        mapeamento = {}
        for campo, variacoes in colunas_esperadas.items():
            for variacao in variacoes:
                coluna_encontrada = None
                for col in df.columns:
                    if str(col).strip().upper() == str(variacao).strip().upper():
                        coluna_encontrada = col
                        break
                if coluna_encontrada:
                    mapeamento[campo] = coluna_encontrada
                    break
        
        if 'descricao' not in mapeamento:
            return jsonify({
                'success': False,
                'message': f'Coluna "Descrição" não encontrada. Colunas: {", ".join(df.columns.tolist())}'
            })
        
        def pegar_valor(row, campo, tipo='str', padrao=None):
            try:
                if campo not in mapeamento:
                    return padrao
                valor = row[mapeamento[campo]]
                if pd.isna(valor) or valor == '':
                    return padrao
                if tipo == 'str':
                    return str(valor).strip() if valor else padrao
                elif tipo == 'float':
                    return float(valor) if valor else (padrao or 0)
                elif tipo == 'int':
                    return int(float(valor)) if valor else (padrao or 0)
                return valor
            except:
                return padrao
        
        importados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                descricao = pegar_valor(row, 'descricao', 'str', '')
                if not descricao:
                    continue
                
                produto = Produto(
                    descricao=descricao,
                    codigo_barras=pegar_valor(row, 'codigo_barras', 'str', None),
                    unidade=pegar_valor(row, 'unidade', 'str', None),
                    marca=pegar_valor(row, 'marca', 'str', None),
                    categoria=pegar_valor(row, 'categoria', 'str', None),
                    fornecedor=pegar_valor(row, 'fornecedor', 'str', None),
                    preco_compra=pegar_valor(row, 'preco_compra', 'float', 0),
                    preco_venda=pegar_valor(row, 'preco_venda', 'float', 0),
                    estoque_minimo=pegar_valor(row, 'estoque_minimo', 'int', 0),
                    estoque_maximo=pegar_valor(row, 'estoque_maximo', 'int', 0),
                    observacao=pegar_valor(row, 'observacao', 'str', None),
                    estoque_atual=0,
                    ativo=True
                )
                
                db.session.add(produto)
                importados += 1
            except Exception as e:
                erros.append(f"Linha {index + 2}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'importados': importados,
            'erros': erros if erros else None,
            'message': f'{importados} produtos importados com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})


# ==================== BACKUP ====================

import shutil
from datetime import datetime

BACKUP_DIR = 'Backup'
AUTO_BACKUP_DIR = os.path.join(BACKUP_DIR, 'AutoBackup')
MANUAL_BACKUP_DIR = os.path.join(BACKUP_DIR, 'ManualBackup')

def criar_diretorios_backup():
    """Criar diretórios de backup"""
    try:
        os.makedirs(AUTO_BACKUP_DIR, exist_ok=True)
        os.makedirs(MANUAL_BACKUP_DIR, exist_ok=True)
    except Exception as e:
        print(f"Erro ao criar diretórios de backup: {e}")

@app.route('/api/backup/criar', methods=['POST'])
def api_backup_criar():
    """Criar backup manual"""
    try:
        criar_diretorios_backup()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"estoque_web_{timestamp}.db"
        backup_path = os.path.join(MANUAL_BACKUP_DIR, backup_filename)
        
        # Copiar banco de dados
        db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        shutil.copy2(db_path, backup_path)
        
        return jsonify({
            'success': True,
            'message': f'Backup criado: {backup_filename}',
            'filename': backup_filename
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao criar backup: {str(e)}'})

@app.route('/api/backup/listar', methods=['GET'])
def api_backup_listar():
    """Listar backups disponíveis"""
    try:
        criar_diretorios_backup()
        
        backups = []
        
        # Listar backups manuais
        if os.path.exists(MANUAL_BACKUP_DIR):
            for arquivo in os.listdir(MANUAL_BACKUP_DIR):
                if arquivo.endswith('.db'):
                    caminho = os.path.join(MANUAL_BACKUP_DIR, arquivo)
                    tamanho = os.path.getsize(caminho)
                    timestamp = os.path.getmtime(caminho)
                    data = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")
                    
                    backups.append({
                        'nome': arquivo,
                        'caminho': caminho,
                        'tamanho': tamanho,
                        'data': data,
                        'tipo': 'Manual'
                    })
        
        # Ordenar por data (mais recente primeiro)
        backups.sort(key=lambda x: x['data'], reverse=True)
        
        return jsonify({'success': True, 'backups': backups})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao listar backups: {str(e)}'})

@app.route('/api/backup/restaurar', methods=['POST'])
def api_backup_restaurar():
    """Restaurar backup"""
    try:
        data = request.json
        backup_filename = data.get('filename')
        
        if not backup_filename:
            return jsonify({'success': False, 'message': 'Arquivo não especificado'})
        
        backup_path = os.path.join(MANUAL_BACKUP_DIR, backup_filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'message': 'Arquivo de backup não encontrado'})
        
        # Criar backup do estado atual antes de restaurar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_atual = f"estoque_web_backup_antes_restauracao_{timestamp}.db"
        backup_atual_path = os.path.join(AUTO_BACKUP_DIR, backup_atual)
        
        db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        shutil.copy2(db_path, backup_atual_path)
        
        # Restaurar backup
        shutil.copy2(backup_path, db_path)
        
        return jsonify({
            'success': True,
            'message': 'Backup restaurado com sucesso. Reinicie o sistema.'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao restaurar backup: {str(e)}'})

@app.route('/api/backup/deletar', methods=['POST'])
def api_backup_deletar():
    """Deletar backup"""
    try:
        data = request.json
        backup_filename = data.get('filename')
        
        if not backup_filename:
            return jsonify({'success': False, 'message': 'Arquivo não especificado'})
        
        backup_path = os.path.join(MANUAL_BACKUP_DIR, backup_filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'message': 'Arquivo de backup não encontrado'})
        
        os.remove(backup_path)
        
        return jsonify({'success': True, 'message': 'Backup deletado com sucesso'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao deletar backup: {str(e)}'})

# ==================== GERENCIAMENTO DE USUÁRIOS ====================

@app.route('/usuarios')
def usuarios():
    """Página de gerenciamento de usuários"""
    if not session.get('usuario_admin'):
        return redirect(url_for('index'))
    return render_template('usuarios.html')

@app.route('/api/usuarios/listar', methods=['GET'])
def api_usuarios_listar():
    """Listar todos os usuários"""
    try:
        if not session.get('usuario_admin'):
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        usuarios = Usuario.query.all()
        lista = [{
            'id': u.id,
            'usuario': u.usuario,
            'nome': u.nome,
            'admin': u.admin,
            'tipo': u.tipo if hasattr(u, 'tipo') else ('master' if u.admin else 'comum'),
            'ativo': u.ativo
        } for u in usuarios]
        
        return jsonify(lista)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/usuarios/criar', methods=['POST'])
def api_usuarios_criar():
    """Criar novo usuário"""
    try:
        if not session.get('usuario_admin'):
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        dados = request.json
        usuario = dados.get('usuario')
        senha = dados.get('senha')
        nome = dados.get('nome')
        tipo = dados.get('tipo', 'comum')  # comum, intermediario, master
        admin = tipo == 'master'  # Master é admin
        
        if not usuario or not senha or not nome:
            return jsonify({'success': False, 'message': 'Preencha todos os campos'})
        
        # Verificar se já existe
        existe = Usuario.query.filter_by(usuario=usuario).first()
        if existe:
            return jsonify({'success': False, 'message': 'Login já existe'})
        
        # Criar novo usuário
        novo_usuario = Usuario(
            usuario=usuario,
            senha=generate_password_hash(senha),
            nome=nome,
            admin=admin,
            tipo=tipo,
            ativo=True
        )
        
        db.session.add(novo_usuario)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Usuário criado com sucesso'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/usuarios/<int:id>/desativar', methods=['POST'])
def api_usuarios_desativar(id):
    """Desativar usuário"""
    try:
        if not session.get('usuario_admin'):
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        usuario = Usuario.query.get(id)
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'})
        
        if usuario.usuario == 'master':
            return jsonify({'success': False, 'message': 'Não é possível desativar o usuário master'})
        
        usuario.ativo = False
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Usuário desativado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/usuarios/<int:id>/ativar', methods=['POST'])
def api_usuarios_ativar(id):
    """Ativar usuário"""
    try:
        if not session.get('usuario_admin'):
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        usuario = Usuario.query.get(id)
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'})
        
        usuario.ativo = True
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Usuário ativado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/usuarios/<int:id>/excluir', methods=['DELETE'])
def api_usuarios_excluir(id):
    """Excluir usuário"""
    try:
        if not session.get('usuario_admin'):
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        usuario = Usuario.query.get(id)
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'})
        
        if usuario.usuario == 'master':
            return jsonify({'success': False, 'message': 'Não é possível excluir o usuário master'})
        
        db.session.delete(usuario)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Usuário excluído'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ==================== APPLICATION STARTUP ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        criar_diretorios_backup()
        optimize_sqlite_once()  # Otimizar SQLite na inicialização
    app.run(debug=False, host='127.0.0.1', port=5000, threaded=True)
